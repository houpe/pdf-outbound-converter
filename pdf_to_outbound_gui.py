#!/usr/bin/env python3.10
"""
PDF出库单转Excel工具 - PySide6桌面版
"""

import pdfplumber
import openpyxl
import re
import os
import sys
import threading
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                                 QHBoxLayout, QLabel, QLineEdit, QPushButton,
                                 QTextBrowser, QGroupBox, QFileDialog, QMessageBox,
                                 QProgressBar, QGridLayout, QFrame)
from PySide6.QtCore import Qt, Signal, QThread, QUrl
from PySide6.QtGui import QFont, QIcon, QDesktopServices


def resource_path(relative_path):
    """Return absolute path to a bundled resource for PyInstaller."""
    if getattr(sys, 'frozen', False):
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)


COLOR = {
    "bg":          "#F0F4F0",
    "card":        "#FFFFFF",
    "card-border": "#D8E6D8",
    "primary":     "#217346",
    "primary-hov": "#1A5E38",
    "primary-press":"#145030",
    "primary-10":  "#E8F5EC",
    "accent":      "#10B981",
    "text":        "#1E293B",
    "text-muted":  "#64748B",
    "input-bg":    "#F8FAF8",
    "input-border":"#CBD5C3",
    "input-focus": "#217346",
    "log-bg":      "#1A1F2E",
    "log-text":    "#A8E6CF",
    "log-dim":     "#6B7280",
    "success":     "#10B981",
    "error":       "#EF4444",
    "progress-bg": "#E2E8E2",
    "shadow":      "rgba(33,115,70,0.08)",
}

QSS_APP = f"""
/* ── Global ─────────────────────────────────────────────── */
QWidget {{
    background-color: {COLOR["bg"]};
    color: {COLOR["text"]};
    font-family: "PingFang SC", "Microsoft YaHei", "Helvetica Neue", Arial, sans-serif;
    font-size: 13px;
}}

QMainWindow {{
    background-color: {COLOR["bg"]};
}}

QLabel {{
    color: {COLOR["text"]};
    font-size: 13px;
    font-weight: 500;
}}

/* ── GroupBox (Card) ────────────────────────────────────── */
QGroupBox {{
    background-color: {COLOR["card"]};
    border: 1px solid {COLOR["card-border"]};
    border-radius: 12px;
    margin-top: 28px;
    padding-top: 20px;
    font-size: 14px;
    font-weight: 600;
    color: {COLOR["primary"]};
}}

QGroupBox::title {{
    subcontrol-origin: margin;
    subcontrol-position: top left;
    left: 16px;
    padding: 6px 16px;
    background-color: {COLOR["primary"]};
    color: #FFFFFF;
    border-radius: 8px 8px 8px 2px;
    font-weight: 700;
    letter-spacing: 0.5px;
}}

/* ── LineEdit ───────────────────────────────────────────── */
QLineEdit {{
    background-color: {COLOR["input-bg"]};
    border: 1.5px solid {COLOR["input-border"]};
    border-radius: 8px;
    padding: 8px 12px;
    font-size: 13px;
    color: {COLOR["text"]};
    selection-background-color: {COLOR["primary-10"]};
    transition: border-color 0.2s ease;
}}

QLineEdit:focus {{
    border: 1.5px solid {COLOR["input-focus"]};
    background-color: #FFFFFF;
}}

QLineEdit:hover {{
    border: 1.5px solid {COLOR["accent"]};
}}

QLineEdit:disabled {{
    background-color: {COLOR["progress-bg"]};
    color: {COLOR["text-muted"]};
}}

/* ── QPushButton (Primary) ─────────────────────────────── */
QPushButton {{
    background-color: {COLOR["primary"]};
    color: #FFFFFF;
    font-size: 14px;
    font-weight: 700;
    padding: 10px 28px;
    border: none;
    border-radius: 10px;
    letter-spacing: 1px;
}}

QPushButton:hover {{
    background-color: {COLOR["primary-hov"]};
}}

QPushButton:pressed {{
    background-color: {COLOR["primary-press"]};
}}

QPushButton:disabled {{
    background-color: {COLOR["input-border"]};
    color: {COLOR["text-muted"]};
}}

/* ── QPushButton (Browse / Secondary) ───────────────────── */
QPushButton#browseBtn {{
    background-color: {COLOR["primary"]};
    color: #FFFFFF;
    font-size: 12px;
    font-weight: 600;
    padding: 7px 0;
    border: none;
    border-radius: 8px;
    min-width: 70px;
}}

QPushButton#browseBtn:hover {{
    background-color: {COLOR["primary-hov"]};
}}

QPushButton#browseBtn:pressed {{
    background-color: {COLOR["primary-press"]};
}}

/* ── QTextEdit (Log Console) ────────────────────────────── */
QTextEdit {{
    background-color: {COLOR["log-bg"]};
    color: {COLOR["log-text"]};
    border: 1px solid #2D3348;
    border-radius: 10px;
    font-family: "JetBrains Mono", "Menlo", "Monaco", "Consolas", "Courier New", monospace;
    font-size: 12px;
    line-height: 1.5;
    padding: 8px 4px;
    selection-background-color: {COLOR["primary"]};
}}

/* ── QProgressBar ───────────────────────────────────────── */
QProgressBar {{
    border: none;
    border-radius: 4px;
    height: 6px;
    background-color: {COLOR["progress-bg"]};
    text-align: center;
}}

QProgressBar::chunk {{
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 {COLOR["primary"]}, stop:1 {COLOR["accent"]});
    border-radius: 4px;
}}
"""


class PathEdit(QLineEdit):
    def __init__(self, placeholder="", parent=None):
        super().__init__(parent)
        self.setPlaceholderText(placeholder)


class HeaderWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedHeight(56)
        self._layout = QHBoxLayout(self)
        self._layout.setContentsMargins(16, 8, 16, 8)

        self.icon_label = QLabel()
        self.icon_label.setFixedSize(36, 36)
        self.icon_label.setStyleSheet(f"""
            QLabel {{
                background-color: {COLOR["primary"]};
                border-radius: 10px;
                color: #FFFFFF;
                font-size: 18px;
                font-weight: 900;
                qproperty-alignment: AlignCenter;
            }}
        """)
        self.icon_label.setText("📄")

        self.title_label = QLabel("PDF出库单 <span style='color:" + COLOR["primary"] + "'>转Excel</span>")
        self.title_label.setTextFormat(Qt.RichText)
        self.title_label.setStyleSheet(f"""
            font-size: 20px;
            font-weight: 800;
            color: {COLOR["text"]};
            letter-spacing: 0.5px;
        """)

        self.subtitle = QLabel("智能解析 · 一键转换")
        self.subtitle.setStyleSheet(f"""
            font-size: 11px;
            color: {COLOR["text-muted"]};
            font-weight: 500;
        """)

        right_layout = QVBoxLayout()
        right_layout.setContentsMargins(10, 0, 0, 0)
        right_layout.setSpacing(2)
        right_layout.addWidget(self.title_label)
        right_layout.addWidget(self.subtitle)

        self._layout.addWidget(self.icon_label)
        self._layout.addLayout(right_layout)
        self._layout.addStretch()


class ConversionWorker(QThread):
    log = Signal(str)
    finished = Signal(bool, str)

    def __init__(self, pdf_path, output_path, merchant_code, template_path):
        super().__init__()
        self.pdf_path = pdf_path
        self.output_path = output_path
        self.merchant_code = merchant_code
        self.template_path = template_path

    def run(self):
        try:
            self.log.emit(f"  正在读取PDF: {self.pdf_path}")
            header_info, items = extract_pdf_data(self.pdf_path)

            self.log.emit("\n  ◆ 提取的头部信息:")
            labels = {
                'order_no': '单据编号', 'receiver_org': '收货机构',
                'supplier_org': '供货机构', 'receiver_name': '收货人',
                'receiver_phone': '收货电话', 'receiver_address': '收货地址',
                'order_date': '订单日期'
            }
            for key, value in header_info.items():
                self.log.emit(f"    {labels.get(key, key)}: {value}")

            self.log.emit(f"\n  ◆ 商品明细 ({len(items)}条):")
            for item in items:
                self.log.emit(f"    {item['item_code']} - {item['item_name']} x {item['quantity']} {item['unit']}")

            self.log.emit("\n  ⚙ 正在生成Excel...")
            create_excel(header_info, items, self.template_path, self.output_path, self.merchant_code)

            self.log.emit(f"\n  ◆ 输出文件:")
            self.log.emit(f"    {self.output_path}")
            self.log.emit(f"\n  💡 共处理 {len(items)} 条记录")
            self.finished.emit(True, self.output_path)
        except Exception as e:
            self.log.emit(f"\n  ✗ 错误: {str(e)}")
            self.finished.emit(False, str(e))


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("PDF出库单转Excel工具")
        self.setMinimumSize(720, 600)
        self.resize(780, 660)

        # 使用资源路径，以支持打包后的可执行文件
        self.template_path = resource_path('OMS出库.xlsx')

        self.init_ui()

    def init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(16, 0, 16, 16)
        main_layout.setSpacing(14)

        header = HeaderWidget(self)
        main_layout.addWidget(header)

        file_group = QGroupBox(" 文件设置 ")
        file_layout = QGridLayout(file_group)
        file_layout.setContentsMargins(18, 18, 18, 18)
        file_layout.setSpacing(12)
        file_layout.setColumnStretch(1, 1)

        pdf_label = QLabel("PDF文件:")
        pdf_label.setStyleSheet(f"color:{COLOR['text-muted']}; font-weight:600; font-size:12px;")
        file_layout.addWidget(pdf_label, 0, 0)

        self.pdf_edit = PathEdit("选择PDF出库单文件...")
        file_layout.addWidget(self.pdf_edit, 0, 1)

        pdf_btn = QPushButton("浏览")
        pdf_btn.setObjectName("browseBtn")
        pdf_btn.setFixedWidth(72)
        pdf_btn.setCursor(Qt.PointingHandCursor)
        pdf_btn.clicked.connect(self.select_pdf)
        file_layout.addWidget(pdf_btn, 0, 2)

        output_label = QLabel("输出路径:")
        output_label.setStyleSheet(f"color:{COLOR['text-muted']}; font-weight:600; font-size:12px;")
        file_layout.addWidget(output_label, 1, 0)

        self.output_edit = PathEdit("自动生成默认文件名...")
        file_layout.addWidget(self.output_edit, 1, 1)

        output_btn = QPushButton("浏览")
        output_btn.setObjectName("browseBtn")
        output_btn.setFixedWidth(72)
        output_btn.setCursor(Qt.PointingHandCursor)
        output_btn.clicked.connect(self.select_output)
        file_layout.addWidget(output_btn, 1, 2)

        merchant_label = QLabel("商户编码:")
        merchant_label.setStyleSheet(f"color:{COLOR['text-muted']}; font-weight:600; font-size:12px;")
        file_layout.addWidget(merchant_label, 2, 0)

        self.merchant_edit = PathEdit("请输入商户编码")
        file_layout.addWidget(self.merchant_edit, 2, 1)

        main_layout.addWidget(file_group)

        btn_layout = QHBoxLayout()
        self.convert_btn = QPushButton("开 始 转 换")
        self.convert_btn.setCursor(Qt.PointingHandCursor)
        self.convert_btn.setMinimumHeight(44)
        self.convert_btn.setMinimumWidth(220)
        self.convert_btn.clicked.connect(self.start_conversion)
        btn_layout.addWidget(self.convert_btn)
        btn_layout.setAlignment(Qt.AlignCenter)
        main_layout.addLayout(btn_layout)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setTextVisible(False)
        self.progress.setMinimumHeight(6)
        main_layout.addWidget(self.progress)

        log_group = QGroupBox(" 转换日志 ")
        log_layout = QVBoxLayout(log_group)
        log_layout.setContentsMargins(14, 14, 14, 14)

        self.log_text = QTextBrowser()
        self.log_text.setReadOnly(True)
        self.log_text.setOpenExternalLinks(False)
        self.log_text.anchorClicked.connect(self.on_link_clicked)
        log_layout.addWidget(self.log_text)

        main_layout.addWidget(log_group)

    def select_pdf(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择PDF文件", "", "PDF files (*.pdf)")
        if path:
            self.pdf_edit.setText(path)
            if not self.output_edit.text():
                base = os.path.splitext(path)[0]
                self.output_edit.setText(f"{base}_转换结果.xlsx")

    def select_output(self):
        path, _ = QFileDialog.getSaveFileName(self, "保存Excel文件", "", "Excel files (*.xlsx)")
        if path:
            self.output_edit.setText(path)

    def start_conversion(self):
        pdf_path = self.pdf_edit.text().strip()
        output_path = self.output_edit.text().strip()
        merchant_code = self.merchant_edit.text().strip()

        if not pdf_path:
            QMessageBox.warning(self, "警告", "请选择PDF文件")
            return
        if not os.path.exists(pdf_path):
            QMessageBox.critical(self, "错误", f"找不到文件: {pdf_path}")
            return
        if not output_path:
            QMessageBox.warning(self, "警告", "请设置输出路径")
            return
        if not os.path.exists(self.template_path):
            QMessageBox.critical(self, "错误", f"找不到模板文件: {self.template_path}")
            return

        self.convert_btn.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setRange(0, 0)
        self.log_text.clear()

        self.worker = ConversionWorker(pdf_path, output_path, merchant_code, self.template_path)
        self.worker.log.connect(self.append_log)
        self.worker.finished.connect(self.on_finished)
        self.worker.start()

    def on_link_clicked(self, url):
        path = url.toString()
        if path.startswith("file://"):
            path = path[7:]
        if os.path.exists(path):
            QDesktopServices.openUrl(QUrl.fromLocalFile(path))

    def append_log(self, text, path=None):
        if path:
            escaped_path = path.replace("'", "&apos;").replace('"', '&quot;')
            text = f'<span style="color:{COLOR["accent"]}; font-weight:700">{text}</span>'
            text += f'<br>    <a href="file://{path}" style="color:#60A5FA; text-decoration:none; font-weight:600">📂 {path}</a>'
            text += f'<br>    <span style="color:{COLOR["text-muted"]}; font-size:11px">点击路径可直接打开文件</span>'
        elif text.startswith("  ✗"):
            text = f'<span style="color:{COLOR["error"]}">{text}</span>'
        elif text.startswith("  ✓"):
            text = f'<span style="color:{COLOR["accent"]}; font-weight:700">{text}</span>'
        elif text.startswith("  ◆"):
            text = f'<span style="color:{COLOR["accent"]}; font-weight:600">{text}</span>'
        elif text.startswith("  ⚙"):
            text = f'<span style="color:#F59E0B; font-weight:600">{text}</span>'
        elif text.startswith("  💡"):
            text = f'<span style="color:#60A5FA; font-weight:600">{text}</span>'
        else:
            text = f'<span style="color:{COLOR["log-dim"]}">{text}</span>'

        self.log_text.append(text)
        self.log_text.verticalScrollBar().setValue(self.log_text.verticalScrollBar().maximum())

    def on_finished(self, success, message):
        self.progress.setVisible(False)
        self.convert_btn.setEnabled(True)
        if success:
            self.append_log("  ✓ 转换完成!", path=message)
            QMessageBox.information(self, "成功", f"转换完成!\n输出文件: {message}\n\n点击日志中的路径可直接打开文件")
        else:
            self.append_log(f"  ✗ 错误: {message}")
            QMessageBox.critical(self, "错误", message)


def extract_pdf_data(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        all_tables = []
        for page in pdf.pages:
            full_text += page.extract_text() or ""
            tables = page.extract_tables()
            all_tables.extend(tables)
    return parse_header(full_text), parse_items(all_tables)


def parse_header(text):
    info = {}
    match = re.search(r'单据编号[：:]\s*(\S+)', text)
    info['order_no'] = match.group(1) if match else ''
    match = re.search(r'收货机构[：:]\s*([^  ]+?)(?=\s+[^\s：:]+[：:]|$)', text)
    info['receiver_org'] = match.group(1).strip() if match else ''
    match = re.search(r'供货机构[：:]\s*([^  ]+?)(?=\s+[^\s：:]+[：:]|$)', text)
    info['supplier_org'] = match.group(1).strip() if match else ''
    match = re.search(r'收货人[：:]\s*([^  ]+?)(?=\s+[^\s：:]+[：:]|$)', text)
    info['receiver_name'] = match.group(1).strip() if match else ''
    match = re.search(r'收货电话[：:]\s*(\d+)', text)
    info['receiver_phone'] = match.group(1) if match else ''
    match = re.search(r'收货地址[：:]\s*(.+?)(?=\n|$)', text)
    info['receiver_address'] = match.group(1).strip() if match else ''
    match = re.search(r'订单日期[：:]\s*(\S+)', text)
    info['order_date'] = match.group(1) if match else ''
    return info


def parse_items(tables):
    items = []
    for table in tables:
        for row in table:
            if not row or len(row) < 6:
                continue
            first_cell = str(row[0]).strip()
            if not first_cell.isdigit():
                continue
            item = {
                'category': str(row[1]).strip() if row[1] else '',
                'item_code': str(row[2]).strip() if row[2] else '',
                'item_name': str(row[3]).strip() if row[3] else '',
                'spec': str(row[4]).strip().replace('\n', '') if row[4] else '',
                'unit': str(row[5]).strip() if row[5] else '',
                'quantity': str(row[6]).strip() if row[6] else '0',
                'remark': str(row[7]).strip() if row[7] else '',
            }
            if item['item_code'] and item['item_name']:
                items.append(item)
    return items


def create_excel(header_info, items, template_path, output_path, merchant_code=''):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            cell.value = None
    receiver_info = f"{header_info.get('receiver_name', '')},{header_info.get('receiver_phone', '')},{header_info.get('receiver_address', '')}"
    for i, item in enumerate(items, start=3):
        ws.cell(row=i, column=1, value=header_info.get('order_no', ''))
        ws.cell(row=i, column=2, value=merchant_code)
        ws.cell(row=i, column=3, value=header_info.get('supplier_org', ''))
        ws.cell(row=i, column=4, value='')
        ws.cell(row=i, column=5, value=receiver_info)
        ws.cell(row=i, column=6, value=item['item_code'])
        ws.cell(row=i, column=7, value='')
        ws.cell(row=i, column=8, value=int(item['quantity']) if item['quantity'].isdigit() else item['quantity'])
        ws.cell(row=i, column=9, value=header_info.get('receiver_org', ''))
        ws.cell(row=i, column=10, value=item['item_name'])
    wb.save(output_path)


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    app.setStyleSheet(QSS_APP)

    base_dir = os.path.dirname(os.path.abspath(__file__))
    icon_path = None
    for ext in ['ico', 'icns', 'png']:
        candidate = os.path.join(base_dir, f'app_icon.{ext}')
        if os.path.exists(candidate):
            icon_path = candidate
            break
    if icon_path:
        app.setWindowIcon(QIcon(icon_path))

    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
