"""
WMS 转换服务 - 文件转换模块
包含文件转换和 Excel 生成逻辑。
"""

import os
import shutil
import uuid
from datetime import date
from typing import Dict, List, Any

from fastapi import HTTPException, UploadFile

from config import (
    MAX_FILE_COUNT,
    MAX_FILE_SIZE,
    DOWNLOADS_DIR,
    UPLOADS_DIR,
    OMS_TEMPLATE,
    TEMPLATES,
)
from database import get_split_map
from strategies import get_strategy


async def _do_convert(
    files: List[UploadFile],
    template_key: str,
    merchant_code: str,
    template_cfg: Dict[str, Any],
) -> Dict[str, Any]:
    """执行文件转换"""
    if template_cfg is None:
        raise HTTPException(status_code=400, detail=f"未知模板: {template_key}")

    if len(files) > MAX_FILE_COUNT:
        raise HTTPException(status_code=400, detail=f"文件数量超限，最多 {MAX_FILE_COUNT} 个")

    if not merchant_code:
        merchant_code = template_cfg["merchant_code"]

    strategy = get_strategy(template_key)

    all_items: List[Dict[str, str]] = []
    header_info: Dict[str, str] = {}
    parsed_files = 0

    for file in files:
        ext = os.path.splitext(file.filename)[1].lower()
        allowed = [a.strip() for a in template_cfg["accept"].split(",")]
        if ext not in allowed:
            raise HTTPException(
                status_code=400,
                detail=f"文件 '{file.filename}' 不支持。模板 '{template_cfg['name']}' 仅支持 {template_cfg['accept']}",
            )

        safe_name = f"{uuid.uuid4().hex}{ext}"
        upload_path = UPLOADS_DIR / safe_name
        file_saved = False

        try:
            total = 0
            with open(upload_path, "wb") as f:
                while True:
                    chunk = await file.read(1024 * 1024)
                    if not chunk:
                        break
                    total += len(chunk)
                    if total > MAX_FILE_SIZE:
                        raise HTTPException(status_code=400, detail=f"文件 '{file.filename}' 超出大小限制 (50MB)")
                    f.write(chunk)
            file_saved = True
        except HTTPException:
            if upload_path.exists():
                upload_path.unlink(missing_ok=True)
            raise
        except Exception as e:
            if upload_path.exists():
                upload_path.unlink(missing_ok=True)
            raise HTTPException(status_code=500, detail=f"文件保存失败: {str(e)}")

        try:
            file_header, items = strategy.parse(str(upload_path), file.filename)

            for item in items:
                item.setdefault("order_no", file_header.get("order_no", ""))
                item.setdefault("receiver_org", file_header.get("receiver_org", ""))
                item.setdefault("receiver_name", file_header.get("receiver_name", ""))
                item.setdefault("receiver_phone", file_header.get("receiver_phone", ""))
                item.setdefault("receiver_address", file_header.get("receiver_address", ""))
                item.setdefault("source_file", file.filename)

            all_items.extend(items)
            if file_header.get("receiver_org"):
                header_info = file_header
            parsed_files += 1
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"文件 '{file.filename}' 转换失败: {str(e)}")
        finally:
            if file_saved and upload_path.exists():
                upload_path.unlink(missing_ok=True)

    if not all_items:
        raise HTTPException(status_code=400, detail="未能解析到任何商品数据")

    today = date.today().strftime("%Y%m%d")
    out_name = f"{template_cfg['name']}_{today}"

    output_filename = f"{out_name}.xlsx"
    output_path = DOWNLOADS_DIR / output_filename
    counter = 1
    while output_path.exists():
        output_filename = f"{out_name}_{counter}.xlsx"
        output_path = DOWNLOADS_DIR / output_filename
        counter += 1

    oms_template_path = str(OMS_TEMPLATE)
    if not os.path.exists(oms_template_path):
        raise HTTPException(status_code=500, detail=f"模板文件不存在: {oms_template_path}")

    try:
        create_excel(header_info, all_items, oms_template_path, str(output_path), merchant_code, template_key)
    except HTTPException:
        raise
    except Exception as e:
        if output_path.exists():
            os.remove(output_path)
        raise HTTPException(status_code=500, detail=f"生成结果文件失败: {str(e)}")

    stores: set = set()
    total_qty = 0
    for it in all_items:
        org = it.get("receiver_org", "")
        if org:
            stores.add(org)
        try:
            total_qty += int(float(it["quantity"]))
        except (ValueError, TypeError):
            pass

    stores_list = sorted(stores)
    result = {
        "success": True,
        "filename": output_filename,
        "download_url": f"/downloads/{output_filename}",
        "item_count": len(all_items),
        "parsed_files": parsed_files,
        "store_count": len(stores_list),
        "total_quantity": total_qty,
        "stores_list": stores_list,
        "warnings": [],
    }
    return result


def create_excel(
    header_info: Dict[str, str],
    items: List[Dict[str, str]],
    template_path: str,
    output_path: str,
    merchant_code: str = "",
    template_key: str = "qzz",
) -> None:
    """复制 OMS 模板，清空第3行及以下，填充转换后的数据"""
    shutil.copy2(template_path, output_path)

    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    max_r = ws.max_row
    if max_r > 2:
        for row in ws.iter_rows(min_row=3, max_row=max_r, min_col=1, max_col=10):
            for cell in row:
                cell.value = None

    strategy = get_strategy(template_key)
    split_map = get_split_map(strategy.warehouse_code)

    if strategy.validate_split_codes:
        strategy.validate_items(items, split_map)

    items = sorted(items, key=lambda x: (x.get("receiver_org", ""), x.get("receiver_name", "")))

    for i, item in enumerate(items, start=3):
        item_receiver_name = item.get("receiver_name", header_info.get("receiver_name", ""))
        item_receiver_phone = item.get("receiver_phone", header_info.get("receiver_phone", ""))
        item_receiver_address = item.get("receiver_address", header_info.get("receiver_address", ""))
        item_receiver_org = item.get("receiver_org", header_info.get("receiver_org", ""))
        item_order_no = item.get("order_no", header_info.get("order_no", ""))

        quantity_val = str(item["quantity"]).strip()
        try:
            quantity_val = int(float(quantity_val))
        except (ValueError, TypeError):
            quantity_val = 0

        item_code = str(item["item_code"]).strip()

        ws.cell(row=i, column=1, value=item_order_no)
        ws.cell(row=i, column=2, value=merchant_code)
        ws.cell(row=i, column=3, value=strategy.warehouse_code)
        ws.cell(row=i, column=4, value="")
        ws.cell(row=i, column=5, value=f"{item_receiver_name},{item_receiver_phone},{item_receiver_address}")
        ws.cell(row=i, column=6, value=item_receiver_org)
        ws.cell(row=i, column=7, value=item["item_name"])
        ws.cell(row=i, column=8, value=item_code)

        col9, col10 = strategy.route_quantity(quantity_val, item_code, split_map)
        ws.cell(row=i, column=9, value=col9)
        ws.cell(row=i, column=10, value=col10)

    wb.save(output_path)
