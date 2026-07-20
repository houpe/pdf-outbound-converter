import re
import openpyxl


def parse_bw_excel(file_path):
    """
    解析霸碗模板
    格式：
    - 行1: 标题
    - 行2: 单号: XXX | 日期：XXX
    - 行3: 人员: XXX | 部门: XXX | 备注：XXX
    - 行4: 收货人:
    - 行5: 收货地址:
    - 行6: 表头（序号|产品编码|品名规格|仓库|批号|包装数量|库存数量|产品编码）
    - 行7+: 数据（隔行分布: 7,9,11...）
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    info = {
        "order_no": "",
        "receiver_org": "",
        "receiver_name": "",
        "receiver_phone": "",
        "receiver_address": "",
    }

    # 从行2提取单号（列1="单号:", 列2=值）
    label1 = ws.cell(row=2, column=1).value or ""
    if '单号' in str(label1):
        info['order_no'] = str(ws.cell(row=2, column=2).value or "").strip()

    # 从行3提取备注/目的地（列6="备注：", 列7=值）
    label6 = ws.cell(row=3, column=6).value or ""
    if '备注' in str(label6):
        remark_text = str(ws.cell(row=3, column=7).value or "").strip()
        # 提取目的地（如"贵阳冷冻"）作为门店名称
        dest = remark_text.split()[0] if remark_text else ""
        info['receiver_org'] = dest

    # 从行5提取收货地址（列1="收货地址:", 列2=值）
    label_addr = ws.cell(row=5, column=1).value or ""
    if '收货地址' in str(label_addr):
        info['receiver_address'] = str(ws.cell(row=5, column=2).value or "").strip()

    # 定位表头行
    header_row = None
    for row_idx in range(1, 15):
        cells = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        cell_text = ' '.join([str(c) for c in cells if c])
        if '产品编码' in cell_text and '品名规格' in cell_text:
            header_row = row_idx
            break

    if not header_row:
        raise ValueError("找不到表头")

    # 查找关键字段列
    item_code_col = None
    item_name_col = None
    quantity_col = None

    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val:
            header_text = str(val).strip()
            if '产品编码' in header_text and item_code_col is None:
                item_code_col = col_idx
            if '品名规格' in header_text:
                item_name_col = col_idx
            if '包装数量' in header_text:
                quantity_col = col_idx

    if not item_code_col or not item_name_col or not quantity_col:
        raise ValueError("找不到必要字段列")

    # 解析数据（隔行读取）
    items = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        if (row_idx - header_row) % 2 == 0:
            continue

        item_code = ws.cell(row=row_idx, column=item_code_col).value
        if item_code is None:
            continue
        item_code = str(item_code).strip()

        item_name = ws.cell(row=row_idx, column=item_name_col).value
        if not item_name:
            continue
        item_name = str(item_name).strip()

        quantity = ws.cell(row=row_idx, column=quantity_col).value
        if quantity is None or quantity == 0:
            continue

        try:
            quantity = int(float(quantity))
            if quantity <= 0:
                continue
        except (ValueError, TypeError):
            continue

        items.append({
            "order_no": info['order_no'],
            "receiver_org": info['receiver_org'],
            "receiver_name": info['receiver_name'],
            "receiver_phone": info['receiver_phone'],
            "receiver_address": info['receiver_address'],
            "item_code": item_code,
            "item_name": item_name,
            "quantity": quantity
        })

    wb.close()

    if not items:
        raise ValueError("未解析到任何商品数据")

    return info, items
