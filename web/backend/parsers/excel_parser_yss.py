"""
WMS 转换服务 Excel 解析器 - 湖南尹三顺
解析 WMS 汇总单发货明细 Excel。
"""

from typing import Dict, List, Tuple

import openpyxl

from parsers.base import _strip_spaces, _normalize_receiver_name


def parse_yss_excel(excel_path: str) -> Tuple[Dict[str, str], List[Dict[str, str]]]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    header_row = _find_header_row(ws)
    if header_row == 0:
        raise ValueError("湖南尹三顺模板格式错误：找不到表头行")

    headers = {}
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=c).value or "").strip()
        if val:
            headers[val] = c

    col_org = _find_col(headers, "收货机构")
    col_hz = _find_col(headers, "配送汇总单号")
    col_ps = _find_col(headers, "配送单号")
    col_line = _find_col(headers, "物品行号")
    col_code = _find_col(headers, "物品编码")
    col_name = _find_col(headers, "物品名称")
    col_spec = _find_col(headers, "规格型号")
    col_qty = _find_col(headers, "发货数量")
    col_recv_name = _find_col(headers, "收货人")
    col_recv_phone = _find_col(headers, "收货电话")
    col_recv_addr = _find_col(headers, "收货地址")

    if not col_code and not col_name:
        raise ValueError("湖南尹三顺模板格式错误：找不到物品编码或物品名称列")

    all_records: List[Dict[str, str]] = []
    data_start = header_row + 1

    for r in range(data_start, ws.max_row + 1):
        org = _get_val(ws, r, col_org)
        item_code = _get_val(ws, r, col_code)
        item_name = _get_val(ws, r, col_name)

        if not org and not item_code and not item_name:
            continue
        if not item_code and not item_name:
            continue

        qty_val = ws.cell(row=r, column=col_qty).value if col_qty else None
        qty = 0
        if qty_val is not None:
            try:
                qty = int(float(qty_val))
            except (ValueError, TypeError):
                pass
        if qty <= 0:
            continue

        recv_name = _get_val(ws, r, col_recv_name)
        recv_phone = _get_val(ws, r, col_recv_phone)
        recv_addr = _get_val(ws, r, col_recv_addr)
        order_no = _get_val(ws, r, col_ps) or _get_val(ws, r, col_hz)

        all_records.append({
            "item_code": item_code,
            "item_name": item_name,
            "quantity": str(qty),
            "spec": _get_val(ws, r, col_spec),
            "category": _get_val(ws, r, _find_col(headers, "物品分类")),
            "remark": "",
            "receiver_org": _strip_spaces(org),
            "receiver_name": _normalize_receiver_name(_strip_spaces(recv_name)),
            "receiver_phone": _strip_spaces(recv_phone),
            "receiver_address": _strip_spaces(recv_addr),
            "order_no": order_no,
        })

    info = {
        "order_no": all_records[0]["order_no"] if all_records else "",
        "receiver_org": "",
        "supplier_org": "",
        "receiver_name": "",
        "receiver_phone": "",
        "receiver_address": "",
        "order_date": "",
    }
    
    # 后处理：检查每个门店是否有收件人和门店一致的订单
    # 如果有，则统一该门店所有订单的单号为"第一个单号_HC"，并将所有原始单号记录到备注
    from collections import defaultdict
    store_records = defaultdict(list)
    
    # 按门店分组
    for record in all_records:
        store = record.get("receiver_org", "")
        if store:
            store_records[store].append(record)
    
    # 处理每个门店
    for store, records in store_records.items():
        # 检查是否有收件人和门店一致的订单
        has_match = False
        first_order_no = None
        
        for record in records:
            receiver_name = record.get("receiver_name", "").strip()
            
            # 简化匹配逻辑：收件人包含"尹三顺自助烤肉屋"或收件人为空
            if not receiver_name or "尹三顺自助烤肉屋" in receiver_name:
                has_match = True
                if first_order_no is None:
                    first_order_no = record.get("order_no", "")
                break
        
        # 如果存在匹配，处理该门店的所有订单
        if has_match and first_order_no:
            # 收集该门店的所有原始单号（去重）
            all_order_nos = []
            for record in records:
                ono = record.get("order_no", "")
                if ono and ono not in all_order_nos:
                    all_order_nos.append(ono)
            
            # 统一单号：第一个单号_HC
            unified_order_no = f"{first_order_no}_HC"
            
            # 更新该门店所有订单
            for record in records:
                record["order_no"] = unified_order_no
                record["remark"] = f"YSS原单号: {','.join(all_order_nos)}"
    
    return info, all_records


def _find_header_row(ws) -> int:
    for r in range(1, min(ws.max_row + 1, 10)):
        v1 = str(ws.cell(row=r, column=1).value or "").strip()
        v6 = str(ws.cell(row=r, column=6).value or "").strip()
        if v1 == "收货机构" and v6 == "物品编码":
            return r
    for r in range(1, min(ws.max_row + 1, 10)):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(row=r, column=c).value or "").strip()
            if v == "收货机构":
                return r
    return 0


def _find_col(headers: dict, *names: str) -> int:
    for name in names:
        if name in headers:
            return headers[name]
    for name in names:
        for h_name, h_idx in headers.items():
            if name in h_name:
                return h_idx
    return 0


def _get_val(ws, row: int, col: int) -> str:
    if not col:
        return ""
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v else ""
