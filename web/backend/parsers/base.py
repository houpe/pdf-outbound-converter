"""
WMS 转换服务解析器基础模块
包含通用解析辅助函数。
"""

import os
import re
from typing import List, Optional, Set, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from config import HEADER_LABEL_PATTERN


def _extract_header_value(text: str, label: str) -> str:
    """从 PDF 文本中提取指定标签的值"""
    pattern = rf"{re.escape(label)}[：:]\s*(.*?)(?=\s*(?:{HEADER_LABEL_PATTERN})[：:]|\n|$)"
    match = re.search(pattern, text)
    return match.group(1).strip() if match else ""


def _extract_shop_name(filename: str) -> str:
    """从文件名中提取店铺名称"""
    raw = os.path.splitext(os.path.basename(filename))[0]
    match = re.match(r"^\d+\.\d+(.+?)(?:[-—_（(]|$)", raw)
    if match:
        return match.group(1).strip()
    return re.sub(r"^[\d./]+", "", raw).strip()


def _find_row_label(ws: Worksheet, texts: Set[str]) -> Optional[int]:
    """在工作表中查找包含指定文本的行"""
    for r in range(1, ws.max_row + 1):
        v = str(ws.cell(row=r, column=1).value or "").strip()
        if v in texts:
            return r
    return None


def search_all_cols(ws: Worksheet, target: str) -> List[Tuple[int, int]]:
    """在工作表中搜索包含指定文本的所有单元格"""
    cells = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and str(v).strip() == target:
                cells.append((r, c))
    return cells


def _normalize_receiver_name(name: str) -> str:
    """如果收件人姓名只有一个字，重复成两个字"""
    if not name:
        return name
    stripped = name.strip()
    if len(stripped) == 1:
        return stripped * 2
    return stripped