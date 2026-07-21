"""
WMS 转换服务 Excel 解析器 - 派乐汉堡
解析派乐汉堡销售出库单 Excel（标准表格格式）。

列结构：
  [1]日期 [2]单据编号 [3]收货客户 [4]客户 [5]单据状态
  [6]物料编码 [7]物料名称 [8]规格型号 [9]库存单位 [10]实发数量 [11]仓库

特点：
  - 「客户」列格式为 "地址（收件人姓名）派乐"，用正则提取括号内的姓名和括号前的地址
  - 「收货客户」是客户编码（PHUN...）
  - 「实发数量」固定路由到 I 列（一级/二级单位），由 strategy 控制
"""

import re
from typing import Dict, List, Tuple

import openpyxl

from parsers.base import _strip_spaces, _normalize_receiver_name


def parse_pl_excel(excel_path: str) -> Tuple[Dict[str, str], List[Dict[str, str]]]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    header_row = _find_header_row(ws)
    if header_row == 0:
        raise ValueError("派乐汉堡模板格式错误：找不到表头行")

    headers = _build_headers(ws, header_row)

    col_order_no = _find_col(headers, "单据编号")
    col_cust_code = _find_col(headers, "收货客户")
    col_customer = _find_col(headers, "客户")
    col_code = _find_col(headers, "物料编码")
    col_name = _find_col(headers, "物料名称")
    col_spec = _find_col(headers, "规格型号")
    col_qty = _find_col(headers, "实发数量")

    if not col_code and not col_name:
        raise ValueError("派乐汉堡模板格式错误：找不到物料编码或物料名称列")

    all_records: List[Dict[str, str]] = []
    data_start = header_row + 1

    for r in range(data_start, ws.max_row + 1):
        item_code = _get_val(ws, r, col_code)
        item_name = _get_val(ws, r, col_name)

        # 空行跳过
        if not item_code and not item_name:
            continue

        # 实发数量
        qty_val = ws.cell(row=r, column=col_qty).value if col_qty else None
        qty = 0
        if qty_val is not None:
            try:
                qty = int(float(qty_val))
            except (ValueError, TypeError):
                pass
        if qty <= 0:
            continue

        # 解析「客户」列：地址（收件人）派乐
        customer = _get_val(ws, r, col_customer)
        recv_name, recv_addr = _parse_customer(customer)

        order_no = _get_val(ws, r, col_order_no)
        # 收货客户编码作为门店/收货机构
        recv_org = _get_val(ws, r, col_cust_code)
        # 按客户全文匹配电话（派乐汉堡电话簿：key=客户完整字符串）
        recv_phone = _lookup_phone(customer)

        all_records.append({
            "item_code": item_code,
            "item_name": item_name,
            "quantity": str(qty),
            "spec": _get_val(ws, r, col_spec),
            "remark": "",
            "receiver_org": _strip_spaces(recv_org),
            "receiver_name": _normalize_receiver_name(_strip_spaces(recv_name)),
            "receiver_phone": _strip_spaces(recv_phone),
            "receiver_address": _strip_spaces(recv_addr),
            "order_no": order_no,
        })

    info = {
        "order_no": all_records[0]["order_no"] if all_records else "",
        "receiver_org": "",
        "supplier_org": "",
        "receiver_name": "",
        "receiver_phone": "",
        "receiver_address": "",
        "order_date": "",
    }

    return info, all_records


def _lookup_phone(customer_code: str) -> str:
    """按客户编码查电话簿（database.customer_phones 表）"""
    if not customer_code:
        return ""
    try:
        from database import get_customer_phone
        return get_customer_phone(customer_code, "pl")
    except Exception:
        return ""


def _parse_customer(customer: str) -> Tuple[str, str]:
    """解析「客户」列：'地址（收件人）派乐' → (收件人姓名, 地址)
    括号内是收件人姓名，括号前是地址。
    """
    if not customer:
        return "", ""
    m = re.search(r"[（(]([^）)]*)[）)]", customer)
    name = m.group(1).strip() if m else ""
    addr = re.sub(r"[（(].*$", "", customer).strip()
    return name, addr


def _find_header_row(ws) -> int:
    for r in range(1, min(ws.max_row + 1, 10)):
        v2 = str(ws.cell(row=r, column=2).value or "").strip()
        v6 = str(ws.cell(row=r, column=6).value or "").strip()
        if v2 == "单据编号" and v6 == "物料编码":
            return r
    # 回退：找含「物料编码」的行
    for r in range(1, min(ws.max_row + 1, 10)):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=r, column=c).value or "").strip() == "物料编码":
                return r
    return 0


def _build_headers(ws, header_row: int) -> dict:
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=c).value or "").strip()
        if val:
            headers[val] = c
    return headers


def _find_col(headers: dict, *names: str) -> int:
    for name in names:
        if name in headers:
            return headers[name]
    for name in names:
        for h_name, h_idx in headers.items():
            if name in h_name:
                return h_idx
    return 0


def _get_val(ws, row: int, col: int) -> str:
    if not col:
        return ""
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v else ""
