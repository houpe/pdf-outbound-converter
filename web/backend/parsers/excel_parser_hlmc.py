"""
WMS 转换服务 Excel 解析器 - 欢乐牧场
解析欢乐牧场 Excel 出库单。
"""

import os
from datetime import date
from typing import Dict, List, Tuple

import openpyxl

from config import HLMC_RECEIVERS
from database import get_hlmc_order
from parsers.base import _normalize_receiver_name

SHOP_ABBREVIATIONS = {
    "金桥": "JQ",
    "银泰": "YT",
    "金银潭": "JYT",
    "金银屯": "JYT",
}


def parse_hlmc_excel(excel_path: str) -> Tuple[Dict[str, str], List[Dict[str, str]]]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    all_records: List[Dict[str, str]] = []
    today_str = date.today().strftime("%y%m%d")

    # Detect columns from row 1
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=1, column=c).value or "").strip()
        if val:
            headers[val] = c

    # Check for "门店名称" in headers → Format B (flat rows)
    # Or check for "下单后结余" → Format A (multi-column shops)
    is_format_b = "门店名称" in headers
    is_format_a = "下单后结余" in headers

    if not is_format_a and not is_format_b:
        raise ValueError("欢乐牧场模板格式错误：找不到\"下单后结余\"列")

    if is_format_b:
        shop_rows = _parse_format_b(ws, headers)
    else:
        shop_rows = _parse_format_a(ws, headers)

    if not shop_rows:
        info = {
            "order_no": "", "receiver_org": "", "supplier_org": "",
            "receiver_name": "", "receiver_phone": "", "receiver_address": "",
            "order_date": "",
        }
        return info, all_records

    _assign_order_numbers(shop_rows, today_str)

    for shop_name, rows in shop_rows.items():
        recv = _match_store(shop_name)
        for r in rows:
            r["receiver_name"] = _normalize_receiver_name(recv.get("name", ""))
            r["receiver_phone"] = recv.get("phone", "")
            r["receiver_address"] = recv.get("address", "")
            r["supplier_org"] = ""
            r["order_date"] = ""
            all_records.append(r)

    info = {
        "order_no": all_records[0]["order_no"] if all_records else "",
        "receiver_org": "",
        "supplier_org": "",
        "receiver_name": "",
        "receiver_phone": "",
        "receiver_address": "",
        "order_date": "",
    }
    return info, all_records


def _find_col(headers: dict, *names: str) -> int:
    for name in names:
        if name in headers:
            return headers[name]
    for name in names:
        for h_name, h_idx in headers.items():
            if name in h_name:
                return h_idx
    return 0


def _get_val(ws, row: int, col: int | None) -> str:
    if not col:
        return ""
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v else ""


def _parse_format_b(ws, headers: dict) -> Dict[str, List[dict]]:
    data_start = 2
    row2_col0 = str(ws.cell(row=2, column=1).value or "").strip()
    if row2_col0 and any(kw in row2_col0 for kw in ("外部单号", "商家单号", "请填写", "*")):
        data_start = 3

    col_shop = _find_col(headers, "门店名称")
    col_code = _find_col(headers, "请填写商品编码", "商家商品编码", "外部商品编码", "商品编码")
    col_sku = _find_col(headers, "SKU名称", "商品名称")
    col_remark = _find_col(headers, "备注")
    col_secondary = _find_col(headers, "请填写二级单位商品数量", "二级单位数量")
    col_min = _find_col(headers, "请填写最小单位商品数量", "最小单位数量")
    col_unit = _find_col(headers, "单位", "库存单位")
    col_spec = _find_col(headers, "规格")

    shop_rows: Dict[str, List[dict]] = {}

    for r in range(data_start, ws.max_row + 1):
        shop_name = _get_val(ws, r, col_shop)
        ext_code = _get_val(ws, r, col_code)
        sku_name = _get_val(ws, r, col_sku)

        if not shop_name:
            continue
        if not ext_code and not sku_name:
            continue

        qty = 0
        sec = ws.cell(row=r, column=col_secondary).value if col_secondary else None
        if sec is not None:
            try:
                v = float(sec)
                if v > 0:
                    qty = int(v)
            except:
                pass
        if qty == 0:
            mn = ws.cell(row=r, column=col_min).value if col_min else None
            if mn is not None:
                try:
                    v = float(mn)
                    if v > 0:
                        qty = int(v)
                except:
                    pass

        if qty <= 0:
            continue

        row_data = {
            "item_code": ext_code,
            "item_name": sku_name,
            "quantity": str(qty),
            "unit": _get_val(ws, r, col_unit),
            "spec": _get_val(ws, r, col_spec),
            "category": "",
            "remark": _get_val(ws, r, col_remark),
            "receiver_org": shop_name,
        }
        shop_rows.setdefault(shop_name, []).append(row_data)

    return shop_rows


def _parse_format_a(ws, headers: dict) -> Dict[str, List[dict]]:
    col_surplus = headers.get("下单后结余", 0)
    col_frozen = headers.get("冻结数量的总和", 0)
    col_unit = _find_col(headers, "单位", "库存单位")
    col_sku = _find_col(headers, "SKU名称", "商品名称")
    col_ext = _find_col(headers, "外部商品编码", "商品编码")
    col_spec = _find_col(headers, "规格")
    col_remark = _find_col(headers, "备注")

    shop_start = col_frozen if col_frozen else col_unit
    shop_cols: Dict[int, str] = {}
    for c in range(shop_start + 1, col_surplus):
        v = str(ws.cell(row=1, column=c).value or "").strip()
        if v:
            shop_cols[c] = v
    if not shop_cols:
        for c in range(1, col_surplus):
            v = str(ws.cell(row=1, column=c).value or "").strip()
            if v:
                shop_cols[c] = v

    shop_rows: Dict[str, List[dict]] = {}

    for r in range(2, ws.max_row + 1):
        sku_name = _get_val(ws, r, col_sku)
        ext_code = _get_val(ws, r, col_ext)
        if not sku_name and not ext_code:
            continue

        for col_idx, shop_name in shop_cols.items():
            qty_val = ws.cell(row=r, column=col_idx).value
            if qty_val is None:
                continue
            try:
                qty = int(float(qty_val))
            except:
                continue
            if qty <= 0:
                continue

            shop_rows.setdefault(shop_name, []).append({
                "item_code": ext_code,
                "item_name": sku_name,
                "quantity": str(qty),
                "unit": _get_val(ws, r, col_unit),
                "spec": _get_val(ws, r, col_spec),
                "category": "",
                "remark": _get_val(ws, r, col_remark),
                "receiver_org": shop_name,
            })

    return shop_rows


def _match_store(shop_name: str) -> Dict[str, str]:
    for key, recv in HLMC_RECEIVERS.items():
        if key in shop_name:
            return recv
    return {}


def _assign_order_numbers(shop_rows: Dict[str, List[dict]], today_str: str):
    for shop_name, rows in shop_rows.items():
        prefix = "XX"
        for key, code in SHOP_ABBREVIATIONS.items():
            if key in shop_name:
                prefix = code
                break

        items_sig = sorted([f"{r['remark']}|{r['item_code']}|{r['quantity']}" for r in rows])
        signature = f"{shop_name}|{today_str}|{'__'.join(items_sig)}"

        order_no = get_hlmc_order(shop_name, today_str, signature, prefix)
        shop_rows[shop_name] = [(r | {"order_no": order_no}) for r in rows]
