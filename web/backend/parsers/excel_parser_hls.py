import re
import openpyxl
import sqlite3
from datetime import date
from config import DB_PATH


def _generate_hls_order_no() -> str:
    """
    生成外部单号：日期(YYYYMMDD) + 3位序号
    例: 20260617001, 20260617002 ...
    每日序号独立递增，用 sqlite 事务保证原子性
    """
    today_str = date.today().strftime("%Y%m%d")
    conn = sqlite3.connect(str(DB_PATH), timeout=10)
    conn.isolation_level = None
    try:
        conn.execute("BEGIN")
        cursor = conn.execute(
            "SELECT last_seq FROM hls_sequences WHERE date_prefix = ?",
            (today_str,)
        )
        row = cursor.fetchone()
        if row is None:
            new_seq = 1
            conn.execute(
                "INSERT INTO hls_sequences (date_prefix, last_seq) VALUES (?, ?)",
                (today_str, new_seq)
            )
        else:
            new_seq = row[0] + 1
            conn.execute(
                "UPDATE hls_sequences SET last_seq = ? WHERE date_prefix = ?",
                (new_seq, today_str)
            )
        conn.execute("COMMIT")
    finally:
        conn.close()
    return f"{today_str}{new_seq:03d}"


def parse_hls_excel(file_path):
    """
    解析湖南联昇模板
    格式：
    - 行1: 标题
    - 行2: 客户名称：XXX   联系电话：姓名 手机号
    - 行3: 制单日期：XXX   收货地址：XXX
    - 行4: 表头（序号|商品名称|单支条码|规格|单位|数量）
    - 行5-12: 数据
    - 行13: 合计
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    order_no = _generate_hls_order_no()
    
    info = {
        "order_no": order_no,
        "receiver_org": "",
        "receiver_name": "欧阳",
        "receiver_phone": "15084853928",
        "receiver_address": "长沙市雨花区万纬物流园1-1-16",
    }
    
    # 定位表头行（第4行）
    header_row = None
    for row_idx in range(1, 10):
        cells = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        cell_text = ' '.join([str(c) for c in cells if c])
        if '商品名称' in cell_text and '单支条码' in cell_text:
            header_row = row_idx
            break
    
    if not header_row:
        raise ValueError("找不到'商品名称'和'单支条码'表头")
    
    # 读取表头
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val:
            headers[col_idx] = str(val).strip()
    
    # 查找关键字段列
    item_code_col = None
    item_name_col = None
    quantity_col = None
    
    for col_idx, header_text in headers.items():
        if '单支条码' in header_text or '条码' in header_text:
            item_code_col = col_idx
        if '商品名称' in header_text or '名称' in header_text:
            item_name_col = col_idx
        if '数量' in header_text:
            quantity_col = col_idx
    
    if not item_code_col or not item_name_col or not quantity_col:
        raise ValueError(f"找不到必要字段列: 单支条码={item_code_col}, 商品名称={item_name_col}, 数量={quantity_col}")
    
    # 解析数据（从表头行+1开始，到合计行前结束）
    items = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        # 检查是否是合计行
        first_cell = ws.cell(row=row_idx, column=1).value
        if first_cell and '合计' in str(first_cell):
            break
        
        # 读取商品编码
        item_code = ws.cell(row=row_idx, column=item_code_col).value
        if item_code is None:
            continue
        item_code = str(item_code).strip()
        
        # 读取商品名称
        item_name = ws.cell(row=row_idx, column=item_name_col).value
        if not item_name:
            continue
        item_name = str(item_name).strip()
        
        # 读取数量
        quantity = ws.cell(row=row_idx, column=quantity_col).value
        if quantity is None or quantity == 0:
            continue
        
        try:
            quantity = int(float(quantity))
            if quantity <= 0:
                continue
        except (ValueError, TypeError):
            continue
        
        items.append({
            "order_no": order_no,
            "receiver_org": info['receiver_org'],
            "receiver_name": info['receiver_name'],
            "receiver_phone": info['receiver_phone'],
            "receiver_address": info['receiver_address'],
            "item_code": item_code,
            "item_name": item_name,
            "quantity": quantity
        })
    
    wb.close()
    
    if not items:
        raise ValueError("未解析到任何商品数据")
    
    return info, items
