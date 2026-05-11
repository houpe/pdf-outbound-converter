"""
WMS 转换服务 Excel 解析器 - 黎明屯铁锅炖
解析黎明屯 Excel 出库单。
"""

from typing import Dict, List, Tuple

import openpyxl

from parsers.base import _extract_shop_name, _find_row_label, _normalize_receiver_name, search_all_cols

LMT_RECEIVER_PHONE = "18888888888"


def parse_lmt_excel(excel_path: str, filename: str = "") -> Tuple[Dict[str, str], List[Dict[str, str]]]:
    """解析黎明屯 Excel 出库单"""
    shop_name = _extract_shop_name(filename) if filename else ""

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    info: Dict[str, str] = {}

    # 提取收货机构
    cells = search_all_cols(ws, "收货机构")
    info["receiver_org"] = ""
    for r, c in cells:
        val = str(ws.cell(row=r, column=c + 1).value or "").strip()
        if val and val != "订货机构":
            info["receiver_org"] = val
            break

    # 提取供货机构
    cells = search_all_cols(ws, "供货机构")
    info["supplier_org"] = ""
    for r, c in cells:
        val = str(ws.cell(row=r, column=c + 1).value or "").strip()
        if val and val != "送货机构":
            info["supplier_org"] = val
            break

    # 提取单据号
    r_no = _find_row_label(ws, {"单据号"})
    info["order_no"] = ""
    if r_no:
        info["order_no"] = str(ws.cell(row=r_no, column=2).value or "").strip()

    # 提取收货人信息
    r_rec = _find_row_label(ws, {"收货人"})
    if r_rec:
        rec_name = str(ws.cell(row=r_rec, column=2).value or "").strip()
        cells = search_all_cols(ws, "收货地址")
        rec_addr = ""
        if cells:
            for rr, cc in cells:
                v = ws.cell(row=rr, column=cc + 1).value
                if v is not None:
                    rec_addr = str(v).strip()
                    break

        info["receiver_name"] = _normalize_receiver_name(rec_name) if rec_name else (shop_name or "")
        info["receiver_phone"] = LMT_RECEIVER_PHONE
        info["receiver_address"] = rec_addr
    else:
        info["receiver_name"] = _normalize_receiver_name(shop_name) if shop_name else ""
        info["receiver_phone"] = LMT_RECEIVER_PHONE
        info["receiver_address"] = ""

    # 提取商品项
    items: List[Dict[str, str]] = []
    header_col = None
    header_row = None
    qty_cols: List[int] = []
    for r in range(1, min(ws.max_row + 1, 6)):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(row=r, column=c).value or "").strip()
            if v in ("物品编码", "商品编码"):
                header_col = c
                header_row = r
            elif v == "发货数量":
                qty_cols.append(c)
        if header_col:
            break

    qty_col = None
    for label in ("发货数量", "订货数量", "接单数量"):
        cells = search_all_cols(ws, label)
        if cells:
            qty_col = cells[0][1]
            break
    if qty_col is None:
        qty_col = header_col + 12 if header_col else 15

    if header_col and header_row:
        for r in range(header_row + 1, ws.max_row + 1):
            c3 = ws.cell(row=r, column=header_col).value
            c4 = ws.cell(row=r, column=header_col + 1).value
            if not c3 or not c4:
                continue
            c3s = str(c3).strip()
            if not c3s or c3s in ("物品编码", "商品编码", "合计") or c3s == "上游单据":
                continue
            item = {
                "category": str(ws.cell(row=r, column=header_col - 1).value or "").strip(),
                "item_code": c3s,
                "item_name": str(c4).strip(),
                "spec": str(ws.cell(row=r, column=header_col + 2).value or "").strip(),
                "unit": "",
                "quantity": str(ws.cell(row=r, column=qty_col).value or "0").strip(),
                "remark": "",
            }
            if item["item_code"] and item["item_name"]:
                try:
                    if float(item["quantity"]) <= 0:
                        continue
                except (ValueError, TypeError):
                    pass
                items.append(item)

    return info, items