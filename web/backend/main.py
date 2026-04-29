#!/usr/bin/env python3
"""
WMS PDF/Excel 转换工具 - FastAPI 后端
将PDF/Excel出库单转换为标准OMS出库Excel格式。
"""

import json
import os
import re
import shutil
import uuid
import time
import logging
from contextlib import asynccontextmanager
from pathlib import Path
from datetime import date, datetime
from collections import defaultdict

import pdfplumber
import openpyxl
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse

BASE_DIR = Path(__file__).resolve().parent
# 模板目录：优先查找 backend/templates/（服务器）或 项目根目录的 templates/（本地开发）
ROOT_TEMPLATES = BASE_DIR.parent.parent / "templates"
LOCAL_TEMPLATES = BASE_DIR / "templates"
if LOCAL_TEMPLATES.is_dir():
    TEMPLATES_DIR = LOCAL_TEMPLATES
elif ROOT_TEMPLATES.is_dir():
    TEMPLATES_DIR = ROOT_TEMPLATES
else:
    TEMPLATES_DIR = ROOT_TEMPLATES  # fallback, will fail if neither exists
UPLOADS_DIR = BASE_DIR / "uploads"
DOWNLOADS_DIR = BASE_DIR / "downloads"

def _find_template(filename: str) -> Path:
    """查找模板文件：优先 templates/ 子目录，回退到 BASE_DIR 同级目录"""
    in_templates = TEMPLATES_DIR / filename
    if in_templates.exists():
        return in_templates
    in_base = BASE_DIR / filename
    if in_base.exists():
        return in_base
    return in_templates  # fallback to original path, will fail if not found

OMS_TEMPLATE = _find_template("OMS出库.xlsx")
SPLIT_TEMPLATE = _find_template("商品拆零模板.xlsx")

UPLOADS_DIR.mkdir(exist_ok=True)
DOWNLOADS_DIR.mkdir(exist_ok=True)

# 转换日志文件
LOG_FILE = BASE_DIR / "conversion_log.jsonl"

LOG_FIELDS = ["template_key", "template_name", "file_names", "file_count",
              "item_count", "store_count", "total_quantity", "stores_list",
              "output_filename", "status", "error"]


def _safe_log(record: dict):
    entry = {k: record.get(k) for k in LOG_FIELDS if k in record}
    entry["timestamp"] = datetime.now().isoformat()
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except Exception:
        pass

# 文件限制
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
MAX_FILE_COUNT = 50

# 下载文件保留时间（秒），默认 24 小时
DOWNLOAD_TTL_SECONDS = 86400

# API 限流：每分钟最多 30 次转换请求
RATE_LIMIT_WINDOW = 60  # 秒
RATE_LIMIT_MAX = 30
_rate_limit_store: dict[str, list[float]] = defaultdict(list)


def _check_rate_limit(client_ip: str) -> None:
    now = time.time()
    window_start = now - RATE_LIMIT_WINDOW
    _rate_limit_store[client_ip] = [
        t for t in _rate_limit_store[client_ip] if t > window_start
    ]
    if len(_rate_limit_store[client_ip]) >= RATE_LIMIT_MAX:
        raise HTTPException(status_code=429, detail="请求过于频繁，请稍后再试")
    _rate_limit_store[client_ip].append(now)


def cleanup_expired_downloads():
    """删除超过 TTL 的下载文件"""
    now = time.time()
    for f in DOWNLOADS_DIR.iterdir():
        if f.is_file() and now - f.stat().st_mtime > DOWNLOAD_TTL_SECONDS:
            f.unlink()


@asynccontextmanager
async def lifespan(app: FastAPI):
    """启动时清理过期下载文件，并定期清理"""
    cleanup_expired_downloads()
    import asyncio
    async def _periodic_cleanup():
        while True:
            await asyncio.sleep(3600)
            try:
                cleanup_expired_downloads()
            except Exception:
                pass
    task = asyncio.create_task(_periodic_cleanup())
    yield
    task.cancel()


app = FastAPI(title="WMS PDF/Excel 转换服务", version="3.4.0", lifespan=lifespan)

TEMPLATES = {
    "qzz": {
        "name": "黔寨寨贵州烙锅",
        "accept": ".pdf",
        "merchant_code": "Q20260427013",
    },
    "lmt": {
        "name": "黎明屯铁锅炖",
        "accept": ".xlsx,.xls",
        "merchant_code": "Q20260427017",
    },
    "hlmc": {
        "name": "欢乐牧场",
        "accept": ".xlsx,.xls",
        "merchant_code": "Q20260427015",
    },
}

# 生产环境 CORS：限定来源，credentials 与 wildcard 不能共存
ALLOWED_ORIGINS = os.environ.get(
    "WMS_CORS_ORIGINS",
    "https://www.houpe.top,http://localhost:5173,http://localhost:3000",
).split(",")
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/api/templates")
def list_templates():
    return {
        "templates": [
            {
                "key": key,
                "name": cfg["name"],
                "accept": cfg["accept"],
                "default_merchant_code": cfg["merchant_code"],
            }
            for key, cfg in TEMPLATES.items()
        ]
    }


@app.post("/api/convert")
async def convert_file(
    files: list[UploadFile] = File(...),
    template_key: str = Form(...),
    merchant_code: str = Form(default=""),
):
    template_cfg = TEMPLATES.get(template_key)
    file_names = [f.filename for f in files]
    log_record = {
        "status": "error",
        "template_key": template_key,
        "template_name": template_cfg["name"] if template_cfg else None,
        "file_names": file_names,
        "file_count": len(files),
    }
    try:
        if merchant_code and len(merchant_code) > 64:
            raise HTTPException(status_code=400, detail="商户编码过长（最多64字符）")
        if merchant_code and not re.match(r'^[A-Za-z0-9_-]+$', merchant_code):
            raise HTTPException(status_code=400, detail="商户编码只能包含字母、数字、下划线和连字符")

        result = await _do_convert(files, template_key, merchant_code, template_cfg)
        log_record["status"] = "success"
        log_record["output_filename"] = result["filename"]
        log_record["item_count"] = result["item_count"]
        log_record["parsed_files"] = result["parsed_files"]
        log_record["store_count"] = result["store_count"]
        log_record["total_quantity"] = result["total_quantity"]
        log_record["stores_list"] = result.get("stores_list", [])
        _safe_log(log_record)
        return result
    except HTTPException as e:
        log_record["error"] = e.detail
        log_record["http_status"] = e.status_code
        _safe_log(log_record)
        raise
    except Exception as e:
        log_record["error"] = str(e)
        _safe_log(log_record)
        raise


async def _do_convert(files, template_key, merchant_code, template_cfg):
    if template_cfg is None:
        raise HTTPException(status_code=400, detail=f"未知模板: {template_key}")

    if len(files) > MAX_FILE_COUNT:
        raise HTTPException(status_code=400, detail=f"文件数量超限，最多 {MAX_FILE_COUNT} 个")

    if not merchant_code:
        merchant_code = template_cfg["merchant_code"]

    all_items = []
    header_info = {}
    parsed_files = 0

    for file in files:
        ext = os.path.splitext(file.filename)[1].lower()
        allowed = [a.strip() for a in template_cfg["accept"].split(",")]
        if ext not in allowed:
            raise HTTPException(
                status_code=400,
                detail=f"文件 '{file.filename}' 不支持。模板 '{template_cfg['name']}' 仅支持 {template_cfg['accept']}",
            )

        safe_name = f"{uuid.uuid4().hex}{ext}"
        upload_path = UPLOADS_DIR / safe_name
        file_saved = False

        # 流式写入，避免整个文件加载到内存
        try:
            total = 0
            with open(upload_path, "wb") as f:
                while True:
                    chunk = await file.read(1024 * 1024)  # 1MB chunks
                    if not chunk:
                        break
                    total += len(chunk)
                    if total > MAX_FILE_SIZE:
                        raise HTTPException(status_code=400, detail=f"文件 '{file.filename}' 超出大小限制 (50MB)")
                    f.write(chunk)
            file_saved = True
        except HTTPException:
            if upload_path.exists():
                upload_path.unlink(missing_ok=True)
            raise
        except Exception as e:
            if upload_path.exists():
                upload_path.unlink(missing_ok=True)
            raise HTTPException(status_code=500, detail=f"文件保存失败: {str(e)}")

        try:
            if template_key == "lmt":
                file_header, items = parse_lmt_excel(str(upload_path), file.filename)
            elif template_key == "hlmc":
                file_header, items = parse_hlmc_excel(str(upload_path))
            else:
                file_header, items = extract_pdf_data(str(upload_path))

            for item in items:
                item.setdefault("order_no", file_header.get("order_no", ""))
                item.setdefault("receiver_org", file_header.get("receiver_org", ""))
                item.setdefault("receiver_name", file_header.get("receiver_name", ""))
                item.setdefault("receiver_phone", file_header.get("receiver_phone", ""))
                item.setdefault("receiver_address", file_header.get("receiver_address", ""))

            all_items.extend(items)
            if file_header.get("receiver_org"):
                header_info = file_header
            parsed_files += 1
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"文件 '{file.filename}' 转换失败: {str(e)}")
        finally:
            if file_saved and upload_path.exists():
                upload_path.unlink(missing_ok=True)

    if not all_items:
        raise HTTPException(status_code=400, detail="未能解析到任何商品数据")

    today = date.today().strftime("%Y%m%d")
    out_name = f"{template_cfg['name']}_{today}"

    output_filename = f"{out_name}.xlsx"
    output_path = DOWNLOADS_DIR / output_filename
    counter = 1
    while output_path.exists():
        output_filename = f"{out_name}_{counter}.xlsx"
        output_path = DOWNLOADS_DIR / output_filename
        counter += 1

    oms_template_path = str(OMS_TEMPLATE)
    if not os.path.exists(oms_template_path):
        raise HTTPException(status_code=500, detail=f"模板文件不存在: {oms_template_path}")

    try:
        create_excel(header_info, all_items, oms_template_path, str(output_path), merchant_code, template_key)
    except Exception as e:
        if output_path.exists():
            os.remove(output_path)
        raise HTTPException(status_code=500, detail=f"生成结果文件失败: {str(e)}")

    stores = set()
    total_qty = 0
    for it in all_items:
        org = it.get("receiver_org", "")
        if org:
            stores.add(org)
        try:
            total_qty += int(float(it["quantity"]))
        except (ValueError, TypeError):
            pass

    stores = sorted(stores)
    result = {
        "success": True,
        "filename": output_filename,
        "download_url": f"/downloads/{output_filename}",
        "item_count": len(all_items),
        "parsed_files": parsed_files,
        "store_count": len(stores),
        "total_quantity": total_qty,
        "stores_list": stores,
    }
    return result


@app.get("/api/logs")
def get_logs(limit: int = 200):
    if not LOG_FILE.exists():
        return {"logs": [], "total": 0}
    entries = []
    with open(LOG_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                entries.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    entries.reverse()
    return {"logs": entries[:limit], "total": len(entries)}


@app.get("/downloads/{filename}")
def download_file(filename: str):
    # 防止路径遍历攻击
    safe = Path(filename).name
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="非法文件名")
    filepath = DOWNLOADS_DIR / safe
    if not filepath.exists() or not filepath.is_file():
        raise HTTPException(status_code=404, detail="文件不存在")
    return FileResponse(
        path=str(filepath),
        filename=safe,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def extract_pdf_data(pdf_path):
    """Open PDF with pdfplumber, extract text+tables, parse header and items."""
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        all_tables = []
        for page in pdf.pages:
            full_text += page.extract_text() or ""
            tables = page.extract_tables()
            all_tables.extend(tables)
    return parse_header(full_text), parse_items(all_tables)


def parse_header(text):
    """Use regex to extract header fields from PDF text."""
    info = {}
    match = re.search(r"单据编号[：:]\s*(\S+)", text)
    info["order_no"] = match.group(1) if match else ""
    match = re.search(r"收货机构[：:]\s*(.+?)(?=\n\s*[^\s：:]+[：:]|$)", text)
    info["receiver_org"] = match.group(1).strip() if match else ""
    match = re.search(r"供货机构[：:]\s*(.+?)(?=\n\s*[^\s：:]+[：:]|$)", text)
    info["supplier_org"] = match.group(1).strip() if match else ""
    match = re.search(r"收货人[：:]\s*(.+?)(?=\n\s*[^\s：:]+[：:]|$)", text)
    info["receiver_name"] = match.group(1).strip() if match else ""
    match = re.search(r"收货电话[：:]\s*(\d+)", text)
    info["receiver_phone"] = match.group(1) if match else ""
    match = re.search(r"收货地址[：:]\s*(.+?)(?=\n|$)", text)
    info["receiver_address"] = match.group(1).strip() if match else ""
    match = re.search(r"订单日期[：:]\s*(\S+)", text)
    info["order_date"] = match.group(1) if match else ""
    return info


def parse_items(tables):
    """Iterate tables, find rows starting with digit, extract item fields."""
    items = []
    for table in tables:
        for row in table:
            if not row or len(row) < 6:
                continue
            first_cell = str(row[0]).strip()
            if not first_cell.isdigit():
                continue
            item = {
                "category": str(row[1]).strip() if row[1] else "",
                "item_code": str(row[2]).strip() if row[2] else "",
                "item_name": str(row[3]).strip() if row[3] else "",
                "spec": str(row[4]).strip().replace("\n", "") if row[4] else "",
                "unit": str(row[5]).strip() if row[5] else "",
                "quantity": str(row[6]).strip() if row[6] else "0",
                "remark": str(row[7]).strip() if row[7] else "",
            }
            if item["item_code"] and item["item_name"]:
                items.append(item)
    return items


def _extract_shop_name(filename):
    raw = os.path.splitext(os.path.basename(filename))[0]
    match = re.match(r"^\d+\.\d+(.+?)(?:[-—_（(]|$)", raw)
    if match:
        return match.group(1).strip()
    return re.sub(r"^[\d./]+", "", raw).strip()


def _find_row_label(ws, texts):
    for r in range(1, ws.max_row + 1):
        v = str(ws.cell(row=r, column=1).value or "").strip()
        if v in texts:
            return r
    return None


def search_all_cols(ws, target):
    cells = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and str(v).strip() == target:
                cells.append((r, c))
    return cells


def parse_lmt_excel(excel_path, filename=""):
    shop_name = _extract_shop_name(filename) if filename else ""

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    info = {}

    cells = search_all_cols(ws, "收货机构")
    info["receiver_org"] = ""
    for r, c in cells:
        val = str(ws.cell(row=r, column=c + 1).value or "").strip()
        if val and val != "订货机构":
            info["receiver_org"] = val
            break

    cells = search_all_cols(ws, "供货机构")
    info["supplier_org"] = ""
    for r, c in cells:
        val = str(ws.cell(row=r, column=c + 1).value or "").strip()
        if val and val != "送货机构":
            info["supplier_org"] = val
            break

    r_no = _find_row_label(ws, {"单据号"})
    info["order_no"] = ""
    if r_no:
        info["order_no"] = str(ws.cell(row=r_no, column=2).value or "").strip()

    r_rec = _find_row_label(ws, {"收货人"})
    if r_rec:
        rec_name = str(ws.cell(row=r_rec, column=2).value or "").strip()
        cells = search_all_cols(ws, "收货电话")
        rec_phone_val = ""
        if cells:
            for rr, cc in cells:
                v = ws.cell(row=rr, column=cc + 1).value
                if v is not None:
                    rec_phone_val = v
                    break
        if isinstance(rec_phone_val, float) and rec_phone_val == int(rec_phone_val):
            rec_phone_val = int(rec_phone_val)
        if isinstance(rec_phone_val, (int, float)):
            rec_phone_val = str(rec_phone_val)

        cells = search_all_cols(ws, "收货地址")
        rec_addr = ""
        if cells:
            for rr, cc in cells:
                v = ws.cell(row=rr, column=cc + 1).value
                if v is not None:
                    rec_addr = str(v).strip()
                    break

        info["receiver_name"] = rec_name if rec_name else (shop_name or "")
        info["receiver_phone"] = str(rec_phone_val or "").strip()
        info["receiver_address"] = rec_addr
    else:
        info["receiver_name"] = shop_name if shop_name else ""
        info["receiver_phone"] = ""
        info["receiver_address"] = ""

    items = []
    header_col = None
    header_row = None
    qty_cols = []
    for r in range(1, min(ws.max_row + 1, 6)):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(row=r, column=c).value or "").strip()
            if v in ("物品编码", "商品编码"):
                header_col = c
                header_row = r
            elif v == "发货数量":
                qty_cols.append(c)
        if header_col:
            break

    qty_col = None
    for label in ("发货数量", "订货数量", "接单数量"):
        cells = search_all_cols(ws, label)
        if cells:
            qty_col = cells[0][1]
            break
    if qty_col is None:
        qty_col = header_col + 12 if header_col else 15

    if header_col and header_row:
        for r in range(header_row + 1, ws.max_row + 1):
            c3 = ws.cell(row=r, column=header_col).value
            c4 = ws.cell(row=r, column=header_col + 1).value
            if not c3 or not c4:
                continue
            c3s = str(c3).strip()
            if not c3s or c3s in ("物品编码", "商品编码", "合计") or c3s == "上游单据":
                continue
            item = {
                "category": str(ws.cell(row=r, column=header_col - 1).value or "").strip(),
                "item_code": c3s,
                "item_name": str(c4).strip(),
                "spec": str(ws.cell(row=r, column=header_col + 2).value or "").strip(),
                "unit": "",
                "quantity": str(ws.cell(row=r, column=qty_col).value or "0").strip(),
                "remark": "",
            }
            if item["item_code"] and item["item_name"]:
                items.append(item)

    return info, items


def parse_hlmc_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    all_records = []

    header_row = ws[1]
    col_frozen = None
    col_unit = None
    col_surplus = None
    col_sku_name = None
    col_ext_code = None
    col_sku_unit = None
    col_sku_spec = None

    for c in range(1, ws.max_column + 1):
        val = str(header_row[c - 1].value or "").strip()
        if val == "冻结数量的总和":
            col_frozen = c
        elif val == "单位":
            col_unit = c
        elif val == "下单后结余":
            col_surplus = c
        elif val in ("SKU名称", "商品名称"):
            col_sku_name = c
        elif val == "外部商品编码":
            col_ext_code = c
        elif val == "商品编码" and not col_ext_code:
            col_ext_code = c
        elif val in ("库存单位", "单位"):
            col_sku_unit = c
        elif val == "规格":
            col_sku_spec = c

    if col_surplus is None:
        raise ValueError("欢乐牧场模板格式错误：找不到\"下单后结余\"列")

    shop_start = col_frozen if col_frozen else (col_unit if col_unit else 0)

    shop_cols = {}
    for c in range(shop_start + 1, col_surplus):
        shop_name = str(header_row[c - 1].value or "").strip()
        if shop_name:
            shop_cols[c] = shop_name

    if not shop_cols:
        for c in range(1, col_surplus):
            v = header_row[c - 1].value
            if v and str(v).strip():
                shop_cols[c] = str(v).strip()

    if not shop_cols:
        raise ValueError("欢乐牧场模板格式错误：找不到店铺列")

    # 欢乐牧场默认收件人信息（可通过环境变量 HLMC_RECEIVERS_JSON 覆盖）
    _hlmc_receivers_env = os.environ.get("HLMC_RECEIVERS_JSON", "")
    if _hlmc_receivers_env:
        try:
            HLMC_RECEIVERS = json.loads(_hlmc_receivers_env)
        except json.JSONDecodeError:
            HLMC_RECEIVERS = {
                "银泰":     {"name": "王先生", "phone": "15289437124", "address": "湖北省武汉市武昌区银泰创意城欢乐牧场"},
                "金银潭":   {"name": "白先生", "phone": "18235064843", "address": "湖北省武汉市东西湖区金银潭永旺欢乐牧场"},
                "金桥":     {"name": "张明",   "phone": "13382067388", "address": "湖北省武汉市江岸区金桥永旺欢乐牧场"},
            }
    else:
        HLMC_RECEIVERS = {
            "银泰":     {"name": "王先生", "phone": "15289437124", "address": "湖北省武汉市武昌区银泰创意城欢乐牧场"},
            "金银潭":   {"name": "白先生", "phone": "18235064843", "address": "湖北省武汉市东西湖区金银潭永旺欢乐牧场"},
            "金桥":     {"name": "张明",   "phone": "13382067388", "address": "湖北省武汉市江岸区金桥永旺欢乐牧场"},
        }

    def _match_store(shop_name):
        for key, recv in HLMC_RECEIVERS.items():
            if key in shop_name:
                return recv
        return None

    r = 2
    while r <= ws.max_row:
        sku_name = ws.cell(row=r, column=col_sku_name).value if col_sku_name else ""
        ext_code = ws.cell(row=r, column=col_ext_code).value if col_ext_code else ""
        sku_unit = str(ws.cell(row=r, column=col_unit).value or "").strip() if col_unit else ""
        sku_spec = str(ws.cell(row=r, column=col_sku_spec).value or "").strip() if col_sku_spec else ""

        for col_idx, shop_name in shop_cols.items():
            qty = ws.cell(row=r, column=col_idx).value
            if qty is not None:
                try:
                    qty_val = float(qty)
                except (ValueError, TypeError):
                    qty_val = 0
                if qty_val > 0:
                    recv = _match_store(shop_name) or {}
                    all_records.append({
                        "item_code": str(ext_code or "").strip(),
                        "item_name": str(sku_name or "").strip(),
                        "quantity": str(int(float(qty))),
                        "unit": sku_unit,
                        "spec": sku_spec,
                        "category": "",
                        "remark": "",
                        "receiver_org": shop_name,
                        "receiver_name": recv.get("name", ""),
                        "receiver_phone": recv.get("phone", ""),
                        "receiver_address": recv.get("address", ""),
                        "order_no": "",
                        "supplier_org": "",
                        "order_date": "",
                    })
        r += 1

    info = {
        "order_no": "",
        "receiver_org": "",
        "supplier_org": "",
        "receiver_name": "",
        "receiver_phone": "",
        "receiver_address": "",
        "order_date": "",
    }
    return info, all_records


def create_excel(header_info, items, template_path, output_path, merchant_code="", template_key="qzz"):
    """shutil.copy2 OMS template, clear row3+, fill from row3 with converted data."""
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    max_r = ws.max_row
    if max_r > 2:
        for row in ws.iter_rows(min_row=3, max_row=max_r, min_col=1, max_col=10):
            for cell in row:
                cell.value = None

    # Load 商品拆零 template: {商品编码 -> 是否拆零}
    split_map = {}
    if SPLIT_TEMPLATE.exists():
        swb = openpyxl.load_workbook(str(SPLIT_TEMPLATE), read_only=True, data_only=True)
        sws = swb.active
        for r in range(2, (sws.max_row or 1) + 1):
            code = sws.cell(row=r, column=1).value
            flag = sws.cell(row=r, column=2).value
            if code:
                split_map[str(code).strip()] = str(flag or "").strip()
        swb.close()

    # 先按门店名称排序，无门店则按收件人排序
    items = sorted(items, key=lambda x: (x.get("receiver_org", ""), x.get("receiver_name", "")))

    for i, item in enumerate(items, start=3):
        item_receiver_name = item.get("receiver_name", header_info.get("receiver_name", ""))
        item_receiver_phone = item.get("receiver_phone", header_info.get("receiver_phone", ""))
        item_receiver_address = item.get("receiver_address", header_info.get("receiver_address", ""))
        item_receiver_org = item.get("receiver_org", header_info.get("receiver_org", ""))
        item_order_no = item.get("order_no", header_info.get("order_no", ""))

        quantity_val = str(item["quantity"]).strip()
        try:
            quantity_val = int(float(quantity_val))
        except (ValueError, TypeError):
            quantity_val = 0

        ws.cell(row=i, column=1, value=item_order_no)
        ws.cell(row=i, column=2, value=merchant_code)
        ws.cell(row=i, column=3, value="ZTOWHHY001")
        ws.cell(row=i, column=4, value="")
        ws.cell(row=i, column=5, value=f"{item_receiver_name},{item_receiver_phone},{item_receiver_address}")
        ws.cell(row=i, column=6, value=item_receiver_org)
        ws.cell(row=i, column=7, value=item["item_name"])
        ws.cell(row=i, column=8, value=item["item_code"])

        # 黔寨寨模板：数量统一放最小单位数量(Col10)
        if template_key == "qzz":
            ws.cell(row=i, column=9, value="")
            ws.cell(row=i, column=10, value=quantity_val)
        else:
            # 其他模板：根据商品拆零模板决定放 Col9(二级单位) 还是 Col10(最小单位)
            item_code = str(item["item_code"]).strip()
            split_flag = split_map.get(item_code, "")
            if split_flag == "否":
                # 不拆零 → 放最小单位数量(Col10)
                ws.cell(row=i, column=9, value="")
                ws.cell(row=i, column=10, value=quantity_val)
            else:
                # 拆零 或 未匹配 → 默认放二级单位数量(Col9)
                ws.cell(row=i, column=9, value=quantity_val)
                ws.cell(row=i, column=10, value="")

    wb.save(output_path)
