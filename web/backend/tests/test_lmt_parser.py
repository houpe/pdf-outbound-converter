import sys
from pathlib import Path

import openpyxl
import pytest

backend_dir = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(backend_dir))

from parsers import parse_lmt_excel


def make_lmt_wb(tmp_path, extra_rows=None):
    fpath = tmp_path / "lmt_gen.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="序号")
    ws.cell(row=1, column=2, value="分类")
    ws.cell(row=1, column=3, value="物品编码")
    ws.cell(row=1, column=4, value="物品名称")
    ws.cell(row=1, column=5, value="规格")
    ws.cell(row=1, column=6, value="单位")
    ws.cell(row=1, column=7, value="发货数量")

    ws.cell(row=3, column=1, value="收货机构")
    ws.cell(row=3, column=2, value="测试门店")
    ws.cell(row=4, column=1, value="收货人")
    ws.cell(row=4, column=2, value="张三")
    ws.cell(row=5, column=1, value="收货电话")
    ws.cell(row=5, column=2, value="138")
    ws.cell(row=6, column=1, value="收货地址")
    ws.cell(row=6, column=2, value="北京路")

    r = 2
    for code, name, qty in extra_rows or []:
        ws.cell(row=r, column=1, value=r - 1)
        ws.cell(row=r, column=2, value="冻品")
        ws.cell(row=r, column=3, value=code)
        ws.cell(row=r, column=4, value=name)
        ws.cell(row=r, column=5, value="标规")
        ws.cell(row=r, column=6, value="盒")
        ws.cell(row=r, column=7, value=qty)
        r += 1

    wb.save(str(fpath))
    wb.close()
    return fpath


class TestLmtParser:

    def test_receiver_info(self, tmp_path):
        fpath = make_lmt_wb(tmp_path, [("CODE01", "商品A", 10)])
        header, items = parse_lmt_excel(str(fpath))
        assert header["receiver_org"] == "测试门店"
        assert header["receiver_name"] == "张三"
        assert header["receiver_phone"] == "18888888888"
        assert header["receiver_address"] == "北京路"

    def test_item_extracted(self, tmp_path):
        fpath = make_lmt_wb(tmp_path, [("C1", "商品X", 5)])
        _, items = parse_lmt_excel(str(fpath))
        assert items[0]["item_code"] == "C1"
        assert items[0]["item_name"] == "商品X"
        assert items[0]["quantity"] == "5"

    def test_skips_non_quantity(self, tmp_path):
        fpath = make_lmt_wb(tmp_path, [("C1", "商品A", "abc")])
        _, items = parse_lmt_excel(str(fpath))
        assert items[0]["quantity"] == "abc"

    def test_float_quantity_stored(self, tmp_path):
        fpath = make_lmt_wb(tmp_path, [("C1", "A", 15.9)])
        _, items = parse_lmt_excel(str(fpath))
        assert items[0]["quantity"] == "15.9"

    def test_single_char_receiver_name_duplicated(self, tmp_path):
        fpath = make_lmt_wb(tmp_path, [("C1", "商品A", 10)])
        wb = openpyxl.load_workbook(str(fpath))
        ws = wb.active
        ws.cell(row=4, column=2, value="张")
        wb.save(str(fpath))
        wb.close()
        header, _ = parse_lmt_excel(str(fpath))
        assert header["receiver_name"] == "张张"

    def test_two_char_receiver_name_unchanged(self, tmp_path):
        fpath = make_lmt_wb(tmp_path, [("C1", "商品A", 10)])
        wb = openpyxl.load_workbook(str(fpath))
        ws = wb.active
        ws.cell(row=4, column=2, value="张三")
        wb.save(str(fpath))
        wb.close()
        header, _ = parse_lmt_excel(str(fpath))
        assert header["receiver_name"] == "张三"

    def test_empty_receiver_name_uses_shop(self, tmp_path):
        fpath = make_lmt_wb(tmp_path, [("C1", "商品A", 10)])
        wb = openpyxl.load_workbook(str(fpath))
        ws = wb.active
        ws.cell(row=4, column=2, value="")
        wb.save(str(fpath))
        wb.close()
        header, _ = parse_lmt_excel(str(fpath), filename="测试门店.xlsx")
        assert header["receiver_name"] == "测试门店"

    def test_skips_zero_quantity(self, tmp_path):
        fpath = make_lmt_wb(tmp_path, [
            ("C1", "商品A", 0),
            ("C2", "商品B", 5),
        ])
        _, items = parse_lmt_excel(str(fpath))
        assert len(items) == 1
        assert items[0]["item_code"] == "C2"

    def test_skips_negative_quantity(self, tmp_path):
        fpath = make_lmt_wb(tmp_path, [("C1", "商品A", -3)])
        _, items = parse_lmt_excel(str(fpath))
        assert len(items) == 0
