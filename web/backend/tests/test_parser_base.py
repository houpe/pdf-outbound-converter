import os
import re
import sys
from pathlib import Path

import openpyxl
import pytest

backend_dir = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(backend_dir))

from parsers.base import _find_row_label, search_all_cols, _extract_header_value


class TestFindRowLabel:

    def _make_ws(self, rows_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        for r_idx, row in enumerate(rows_data, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        return ws

    def test_found(self):
        ws = self._make_ws([["Header1"], ["收货机构"], ["Data"]])
        assert _find_row_label(ws, {"收货机构"}) == 2

    def test_not_found(self):
        ws = self._make_ws([["Header1"], ["Data"]])
        assert _find_row_label(ws, {"收货机构"}) is None

    def test_returns_first_match(self):
        ws = self._make_ws([["收货机构"], ["Data"], ["收货机构"]])
        assert _find_row_label(ws, {"收货机构"}) == 1

    def test_empty_worksheet(self):
        ws = openpyxl.Workbook().active
        assert _find_row_label(ws, {"anything"}) is None

    def test_multiple_labels_any_match(self):
        ws = self._make_ws([["Header"], ["供货机构"], ["收货机构"]])
        assert _find_row_label(ws, {"收货机构", "供货机构"}) == 2


class TestSearchAllCols:

    def _make_ws(self, rows_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        for r_idx, row in enumerate(rows_data, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        return ws

    def test_single_match(self):
        ws = self._make_ws([["A", "B"], ["C", "D"]])
        assert search_all_cols(ws, "B") == [(1, 2)]

    def test_no_match(self):
        ws = self._make_ws([["A", "B"]])
        assert search_all_cols(ws, "Z") == []

    def test_multiple_matches(self):
        ws = self._make_ws([["X", "Y", "X"], ["Z"]])
        result = search_all_cols(ws, "X")
        assert (1, 1) in result
        assert (1, 3) in result
        assert len(result) == 2

    def test_case_sensitive(self):
        ws = self._make_ws([["Code"], ["code"]])
        assert search_all_cols(ws, "Code") == [(1, 1)]
        assert search_all_cols(ws, "code") == [(2, 1)]


class TestExtractShopName:

    def _extract(self, name):
        from parsers.base import _extract_shop_name
        return _extract_shop_name(name)

    def test_with_date_and_dash(self):
        assert self._extract("12.25门店名-发货单.xlsx") == "门店名"

    def test_no_date_prefix(self):
        result = self._extract("欢乐牧场-发货单.xlsx")
        assert "欢乐牧场" in result

    def test_unicode_store_name(self):
        result = self._extract("12.25欢乐店-发货单.xlsx")
        assert "欢乐店" in result

    def test_with_parentheses(self):
        result = self._extract("12.25门店A(分店).xlsx")
        assert "门店A" in result


class TestExtractHeaderValue:

    def test_basic_extraction(self):
        text = "单据编号：PS2512210002001\n收货机构：测试门店"
        assert _extract_header_value(text, "单据编号") == "PS2512210002001"

    def test_missing_label(self):
        text = "单据编号：PS001"
        assert _extract_header_value(text, "不存在字段") == ""

    def test_english_colon(self):
        text = "收货机构:测试门店A"
        assert _extract_header_value(text, "收货机构") == "测试门店A"

    def test_multiline(self):
        text = "单据编号：PS001\n收货机构：测试门店\n供货机构：中央厨房"
        assert _extract_header_value(text, "收货机构") == "测试门店"

    def test_empty_text(self):
        assert _extract_header_value("", "单据编号") == ""


class TestLmtParserWithFilename:

    def test_basic_lmt_with_filename(self, tmp_path):
        from parsers import parse_lmt_excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="序号")
        ws.cell(row=1, column=2, value="分类")
        ws.cell(row=1, column=3, value="物品编码")
        ws.cell(row=1, column=4, value="物品名称")
        ws.cell(row=1, column=5, value="规格")
        ws.cell(row=1, column=6, value="单位")
        ws.cell(row=1, column=7, value="发货数量")
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value="冻品")
        ws.cell(row=2, column=3, value="C1")
        ws.cell(row=2, column=4, value="商品")
        ws.cell(row=2, column=7, value=5)
        ws.cell(row=3, column=1, value="收货机构")
        ws.cell(row=3, column=2, value="门店")
        fpath = tmp_path / "lmt_fn.xlsx"
        wb.save(str(fpath))
        wb.close()

        header, items = parse_lmt_excel(str(fpath), filename="test.xlsx")
        assert isinstance(header, dict)
        assert "receiver_org" in header

    def test_empty_filename_no_crash(self, tmp_path):
        from parsers import parse_lmt_excel
        wb = openpyxl.Workbook()
        ws = wb.active
        fpath = tmp_path / "minimal.xlsx"
        wb.save(str(fpath))
        wb.close()

        header, items = parse_lmt_excel(str(fpath), filename="")
        assert isinstance(header, dict)
        assert isinstance(items, list)
