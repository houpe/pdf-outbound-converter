"""
Pytest fixtures for backend tests.
"""

import os
import sys
import sqlite3
import tempfile
from pathlib import Path
from typing import Generator

import openpyxl
import pytest

# Add backend to path for imports
backend_dir = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(backend_dir))


@pytest.fixture
def tmp_db(tmp_path: Path) -> Generator[Path, None, None]:
    """
    Creates a temporary SQLite database for testing.
    Yields the path to the database file.
    Cleans up after the test.
    """
    db_path = tmp_path / "test_split_codes.db"
    conn = sqlite3.connect(str(db_path))
    conn.execute("""
        CREATE TABLE IF NOT EXISTS split_codes (
            code TEXT PRIMARY KEY COLLATE NOCASE,
            split TEXT NOT NULL DEFAULT '是',
            item_name TEXT,
            created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
        )
    """)
    conn.commit()
    conn.close()
    yield db_path
    # Cleanup
    if db_path.exists():
        db_path.unlink()


@pytest.fixture
def test_files_dir(tmp_path: Path) -> Path:
    """
    Creates a directory with test fixture files.
    Returns the path to the directory.
    """
    fixtures_dir = tmp_path / "fixtures"
    fixtures_dir.mkdir(exist_ok=True)
    return fixtures_dir


@pytest.fixture
def lmt_excel(test_files_dir: Path) -> Path:
    """
    Creates a minimal LMT (黎明屯铁锅炖) template Excel file.
    The structure matches what parse_lmt_excel expects:
    - Header info in early rows
    - "物品编码" found in rows 1-6 (parser searches this range)
    - Quantity column via "发货数量" label
    """
    file_path = test_files_dir / "test_lmt.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "配送发货单"

    # Item table header at row 1-6 range (parser searches this range)
    # Row 1: header labels
    ws.cell(row=1, column=1, value="序号")
    ws.cell(row=1, column=2, value="分类")
    ws.cell(row=1, column=3, value="物品编码")  # Parser looks for this in rows 1-6
    ws.cell(row=1, column=4, value="物品名称")
    ws.cell(row=1, column=5, value="规格")
    ws.cell(row=1, column=6, value="单位")
    ws.cell(row=1, column=7, value="发货数量")  # Parser searches for this

    # Header info below item table header (row 3-6)
    ws.cell(row=3, column=1, value="收货机构")
    ws.cell(row=3, column=2, value="测试门店A")
    ws.cell(row=4, column=1, value="供货机构")
    ws.cell(row=4, column=2, value="中央厨房")
    ws.cell(row=5, column=1, value="单据号")
    ws.cell(row=5, column=2, value="PS2512210002001")
    ws.cell(row=6, column=1, value="收货人")
    ws.cell(row=6, column=2, value="张三")
    ws.cell(row=7, column=1, value="收货电话")
    ws.cell(row=7, column=2, value="13800138000")
    ws.cell(row=8, column=1, value="收货地址")
    ws.cell(row=8, column=2, value="北京市朝阳区测试路1号")

    # Item data starting from row 2 (after header row 1)
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="冻品")
    ws.cell(row=2, column=3, value="CODE001")
    ws.cell(row=2, column=4, value="测试商品A")
    ws.cell(row=2, column=5, value="500g/包")
    ws.cell(row=2, column=6, value="包")
    ws.cell(row=2, column=7, value=10)

    # Additional items in later rows (row 10+ to not overlap with header info)
    ws.cell(row=10, column=1, value=2)
    ws.cell(row=10, column=2, value="冻品")
    ws.cell(row=10, column=3, value="CODE002")
    ws.cell(row=10, column=4, value="测试商品B")
    ws.cell(row=10, column=5, value="1kg/袋")
    ws.cell(row=10, column=6, value="袋")
    ws.cell(row=10, column=7, value=5)

    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture
def hlmc_excel(test_files_dir: Path) -> Path:
    """
    Creates a minimal HLMC (欢乐牧场) template Excel file.
    """
    file_path = test_files_dir / "test_hlmc.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    # Header row with required columns
    ws.cell(row=1, column=1, value="SKU名称")
    ws.cell(row=1, column=2, value="外部商品编码")
    ws.cell(row=1, column=3, value="规格")
    ws.cell(row=1, column=4, value="冻结数量的总和")
    ws.cell(row=1, column=5, value="单位")
    ws.cell(row=1, column=6, value="银泰")  # Store 1
    ws.cell(row=1, column=7, value="金银潭")  # Store 2
    ws.cell(row=1, column=8, value="下单后结余")  # End marker

    # Data rows
    ws.cell(row=2, column=1, value="牛肉片")
    ws.cell(row=2, column=2, value="SKU001")
    ws.cell(row=2, column=3, value="500g")
    ws.cell(row=2, column=4, value=0)
    ws.cell(row=2, column=5, value="盒")
    ws.cell(row=2, column=6, value=10)  # Quantity for Store 1
    ws.cell(row=2, column=7, value=5)   # Quantity for Store 2
    ws.cell(row=2, column=8, value=100)

    ws.cell(row=3, column=1, value="羊肉卷")
    ws.cell(row=3, column=2, value="SKU002")
    ws.cell(row=3, column=3, value="300g")
    ws.cell(row=3, column=4, value=0)
    ws.cell(row=3, column=5, value="盒")
    ws.cell(row=3, column=6, value=8)
    ws.cell(row=3, column=7, value=3)
    ws.cell(row=3, column=8, value=50)

    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture
def oms_template(test_files_dir: Path) -> Path:
    """
    Creates a minimal OMS output template Excel file.
    """
    file_path = test_files_dir / "OMS出库.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    # Header row
    ws.cell(row=1, column=1, value="单据编号")
    ws.cell(row=1, column=2, value="商户编码")
    ws.cell(row=1, column=3, value="仓库编码")
    ws.cell(row=1, column=4, value="备注")
    ws.cell(row=1, column=5, value="收货人信息")
    ws.cell(row=1, column=6, value="收货机构")
    ws.cell(row=1, column=7, value="商品名称")
    ws.cell(row=1, column=8, value="商品编码")
    ws.cell(row=1, column=9, value="二级单位数量")
    ws.cell(row=1, column=10, value="最小单位数量")

    # Empty row 2 for template
    ws.cell(row=2, column=1, value="")

    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture
def split_template(test_files_dir: Path) -> Path:
    """
    Creates a minimal split codes template Excel file.
    """
    file_path = test_files_dir / "商品拆零模板.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    # Header row
    ws.cell(row=1, column=1, value="商品编码")
    ws.cell(row=1, column=2, value="是否拆零")

    # Sample data
    ws.cell(row=2, column=1, value="CODE001")
    ws.cell(row=2, column=2, value="是")

    ws.cell(row=3, column=1, value="CODE002")
    ws.cell(row=3, column=2, value="否")

    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture
def mock_main_env(tmp_db: Path, oms_template: Path, split_template: Path, monkeypatch) -> dict:
    """
    Sets up a mock environment for testing main.py endpoints.
    Patches global paths to use temporary files.
    """
    # We'll return paths that tests can use for patching
    return {
        "db_path": tmp_db,
        "oms_template": oms_template,
        "split_template": split_template,
    }