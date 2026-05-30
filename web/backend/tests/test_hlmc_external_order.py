"""
Tests for HLMC (欢乐牧场) external order number assignment logic.
Covers SQLite-based sequence and history management.
"""

import os
import sys
from datetime import date
from pathlib import Path

import openpyxl
import pytest

# Add backend to path
backend_dir = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(backend_dir))

import database
import parsers.excel_parser_hlmc as hlmc_parser
from database import get_db, init_db as init_db_real

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def hlmc_order_db(tmp_path: Path):
    """Create an isolated SQLite database for HLMC order number tests."""
    db_path = tmp_path / "test_hlmc_orders.db"
    conn = database.sqlite3.connect(str(db_path))
    conn.execute("""
        CREATE TABLE IF NOT EXISTS split_codes (
            code TEXT PRIMARY KEY COLLATE NOCASE,
            split TEXT NOT NULL DEFAULT '是',
            item_name TEXT,
            created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS hlmc_sequences (
            date TEXT NOT NULL,
            prefix TEXT NOT NULL,
            seq INTEGER NOT NULL DEFAULT 1,
            PRIMARY KEY (date, prefix)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS hlmc_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shop_name TEXT,
            date TEXT NOT NULL,
            signature TEXT NOT NULL UNIQUE,
            order_no TEXT NOT NULL
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_hlmc_hist_sig ON hlmc_history(signature)")
    conn.commit()
    conn.close()
    yield db_path
    # Cleanup
    if db_path.exists():
        db_path.unlink()


@pytest.fixture
def isolated_hlmc_parser(hlmc_order_db: Path, monkeypatch):
    """
    Monkey-patch database module to use isolated test DB.
    Replaces get_db() and get_hlmc_order() to point to test database.
    """
    def get_test_db():
        conn = database.sqlite3.connect(str(hlmc_order_db))
        conn.row_factory = database.sqlite3.Row
        return conn

    monkeypatch.setattr(database, "get_db", get_test_db)
    
    # Also ensure the parser uses the patched get_db
    monkeypatch.setattr(database, "DB_PATH", hlmc_order_db)
    monkeypatch.setattr(hlmc_parser, "get_hlmc_order", database.get_hlmc_order)
    
    yield hlmc_order_db


@pytest.fixture
def make_hlmc_workbook(tmp_path: Path):
    """
    Factory to create HLMC Format A workbooks with custom store data.
    Returns the file path.
    """
    def _make(stores: dict, file_name: str = "test_store_data.xlsx") -> Path:
        """
        stores: { "store_name": [("item_code", "quantity"), ...] }
        """
        file_path = tmp_path / file_name
        wb = openpyxl.Workbook()
        ws = wb.active

        # Header row
        ws.cell(row=1, column=1, value="SKU名称")
        ws.cell(row=1, column=2, value="外部商品编码")
        ws.cell(row=1, column=3, value="规格")
        ws.cell(row=1, column=4, value="冻结数量的总和")
        ws.cell(row=1, column=5, value="单位")
        
        # Determine store column positions
        col_idx = 6
        store_cols = {}
        for store_name in stores:
            ws.cell(row=1, column=col_idx, value=store_name)
            store_cols[store_name] = col_idx
            col_idx += 1
            
        # End marker
        ws.cell(row=1, column=col_idx, value="下单后结余")
        surplus_col = col_idx
        
        # Data rows
        row_idx = 2
        items = {}
        for store_name, products in stores.items():
            for code, qty in products:
                if code not in items:
                    items[code] = {}
                items[code][store_name] = qty

        for code, store_qtys in items.items():
            ws.cell(row=row_idx, column=1, value=f"商品_{code}")
            ws.cell(row=row_idx, column=2, value=code)
            ws.cell(row=row_idx, column=3, value="标准规格")
            ws.cell(row=row_idx, column=4, value=0)
            ws.cell(row=row_idx, column=5, value="盒")
            
            for store_name, qty in store_qtys.items():
                c = store_cols[store_name]
                ws.cell(row=row_idx, column=c, value=qty)
                
            ws.cell(row=row_idx, column=surplus_col, value=100)
            row_idx += 1
            
        wb.save(str(file_path))
        wb.close()
        return file_path
    
    return _make


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestHlmcSequenceLogic:
    """Test raw sequence generation logic via get_hlmc_order."""
    
    def test_first_assignment_generates_0001(self, isolated_hlmc_parser, monkeypatch):
        """First call for a store/date should generate ...0001."""
        result = database.get_hlmc_order("银泰", "260502", "sig_yintai_v1", "YT")
        assert result == "YT2605020001"

    def test_same_signature_reuses_order_no(self, isolated_hlmc_parser):
        """Same signature on same day should return identical order_no."""
        sig = "金桥|260502|codeA|5"
        no1 = database.get_hlmc_order("金桥", "260502", sig, "JQ")
        no2 = database.get_hlmc_order("金桥", "260502", sig, "JQ")
        assert no1 == no2
        assert no1 == "JQ2605020001"

    def test_different_signature_increments(self, isolated_hlmc_parser):
        """Different signature on same day should increment sequence."""
        sig1 = "银泰|260502|goods_A|10"
        sig2 = "银泰|260502|goods_B|20"
        
        order1 = database.get_hlmc_order("银泰", "260502", sig1, "YT")
        order2 = database.get_hlmc_order("银泰", "260502", sig2, "YT")
        
        assert order1 == "YT2605020001"
        assert order2 == "YT2605020002"

    def test_date_change_resets_count(self, isolated_hlmc_parser):
        sig_day1 = "金银潭|260501|same_goods"
        sig_day2 = "金银潭|260502|same_goods"
        
        order_day1 = database.get_hlmc_order("金银潭", "260501", sig_day1, "JYT")
        order_day2 = database.get_hlmc_order("金银潭", "260502", sig_day2, "JYT")
        
        assert order_day1 == "JYT2605010001"
        assert order_day2 == "JYT2605020001"

    def test_different_prefixes_independent(self, isolated_hlmc_parser):
        """Different store prefixes maintain independent counters."""
        sig_yt = "银泰|260502|item1"
        sig_jq = "金桥|260502|item1"
        
        yt_order = database.get_hlmc_order("银泰", "260502", sig_yt, "YT")
        jq_order1 = database.get_hlmc_order("金桥", "260502", sig_jq, "JQ")
        jq_order2 = database.get_hlmc_order("金桥", "260502", sig_jq, "JQ")
        
        assert yt_order == "YT2605020001"
        assert jq_order1 == "JQ2605020001"
        assert jq_order2 == "JQ2605020001"  # Reuse works


class TestHlmcParserIntegration:
    """Test full parser workflow with SQLite storage."""

    def test_single_file_assigns_correct_prefixes(self, isolated_hlmc_parser, make_hlmc_workbook):
        """One file with multiple stores gets correct prefix per store."""
        data = {
            "银泰店": [("CODE01", 10), ("CODE02", 5)],
            "金桥总店": [("CODE01", 20)],
            "金银潭分店": [("CODE03", 15)],
        }
        fpath = make_hlmc_workbook(data, "multi_store_test.xlsx")
        
        header, items = hlmc_parser.parse_hlmc_excel(str(fpath))
        
        orders_by_store = {}
        for item in items:
            org = item["receiver_org"]
            no = item["order_no"]
            if org not in orders_by_store:
                orders_by_store[org] = set()
            orders_by_store[org].add(no)
            
        # Each store gets 0001 on first import
        for org, order_nos in orders_by_store.items():
            assert len(order_nos) == 1
            order_no = next(iter(order_nos))
            assert order_no.endswith("0001")

            # Verify prefix mapping
            if "银泰" in org:
                assert order_no.startswith("YT")
            elif "金桥" in org:
                assert order_no.startswith("JQ")
            elif "金银潭" in org:
                assert order_no.startswith("JYT")

    def test_import_twice_reuses_order_number(self, isolated_hlmc_parser, make_hlmc_workbook):
        """
        Importing the SAME data twice should reuse the existing order number.
        Tests the signature deduplication logic end-to-end.
        """
        data = {"测试门店": [("GOODS_A", 10), ("GOODS_B", 20)]}
        fpath = make_hlmc_workbook(data, "reuse_test.xlsx")
        
        _, items1 = hlmc_parser.parse_hlmc_excel(str(fpath))
        order_no_1 = items1[0]["order_no"]
        
        # Parse same file again
        _, items2 = hlmc_parser.parse_hlmc_excel(str(fpath))
        order_no_2 = items2[0]["order_no"]
        
        assert order_no_1 == order_no_2
        assert order_no_1.endswith("0001")

    def test_different_data_increments(self, isolated_hlmc_parser, make_hlmc_workbook):
        """
        Importing DIFFERENT data for same store should increment order number.
        """
        # First import: GOODS_A
        fpath1 = make_hlmc_workbook({"门店A": [("GOODS_A", 10)]}, "diff_data_1.xlsx")
        _, items1 = hlmc_parser.parse_hlmc_excel(str(fpath1))
        order_no_1 = items1[0]["order_no"]
        
        # Second import: GOODS_B (different signature)
        fpath2 = make_hlmc_workbook({"门店A": [("GOODS_B", 20)]}, "diff_data_2.xlsx")
        _, items2 = hlmc_parser.parse_hlmc_excel(str(fpath2))
        order_no_2 = items2[0]["order_no"]
        
        assert order_no_1.endswith("0001")
        assert order_no_2.endswith("0002")
        assert order_no_1 != order_no_2

    def test_mixed_reuse_and_increment(self, isolated_hlmc_parser, make_hlmc_workbook):
        """
        Store A imports same data -> reuse.
        Store B imports different data -> increment.
        """
        # Initial state
        fpath1 = make_hlmc_workbook({
            "银泰店": [("ITEM_X", 5)],
            "金桥店": [("ITEM_Y", 10)]
        }, "mixed_initial.xlsx")
        
        _, items1 = hlmc_parser.parse_hlmc_excel(str(fpath1))
        yt_no_1 = next((i["order_no"] for i in items1 if "银泰" in i["receiver_org"]), None)
        jq_no_1 = next((i["order_no"] for i in items1 if "金桥" in i["receiver_org"]), None)
        
        assert yt_no_1.endswith("0001")
        assert jq_no_1.endswith("0001")
        
        # Second batch: 银泰 same data, 金桥 different data
        fpath2 = make_hlmc_workbook({
            "银泰店": [("ITEM_X", 5)],       # Same signature
            "金桥店": [("ITEM_Z", 30)]       # Different signature
        }, "mixed_second.xlsx")
        
        _, items2 = hlmc_parser.parse_hlmc_excel(str(fpath2))
        yt_no_2 = next((i["order_no"] for i in items2 if "银泰" in i["receiver_org"]), None)
        jq_no_2 = next((i["order_no"] for i in items2 if "金桥" in i["receiver_org"]), None)
        
        # 银泰 should reuse 0001
        assert yt_no_2 == yt_no_1
        # 金桥 should increment to 0002
        assert jq_no_2.endswith("0002")
        assert jq_no_2 != jq_no_1

    def test_zero_quantities_skipped(self, isolated_hlmc_parser, make_hlmc_workbook):
        """Rows with zero quantity should not generate order numbers."""
        data = {"测试门店": [("GOODS_A", 0), ("GOODS_B", 0)]}
        fpath = make_hlmc_workbook(data, "zero_qty.xlsx")
        
        header, items = hlmc_parser.parse_hlmc_excel(str(fpath))
        assert len(items) == 0

    def test_missing_surplus_column_raises(self, isolated_hlmc_parser, tmp_path):
        """File without required columns should raise ValueError."""
        fpath = tmp_path / "invalid.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="SKU名称")
        ws.cell(row=1, column=2, value="外部商品编码")
        wb.save(str(fpath))
        wb.close()
        
        with pytest.raises(ValueError, match="找不到.*下单后结余"):
            hlmc_parser.parse_hlmc_excel(str(fpath))


class TestHlmcFormatB:
    """Format B (flat rows with 门店名称 column) parsing + external order integration."""

    @pytest.fixture
    def make_format_b_workbook(self, tmp_path: Path):
        def _make(data: dict, file_name: str = "fmt_b.xlsx") -> Path:
            file_path = tmp_path / file_name
            wb = openpyxl.Workbook()
            ws = wb.active

            ws.cell(row=1, column=1, value="门店名称")
            ws.cell(row=1, column=2, value="SKU名称")
            ws.cell(row=1, column=3, value="外部商品编码")
            ws.cell(row=1, column=4, value="备注")
            ws.cell(row=1, column=5, value="规格")
            ws.cell(row=1, column=6, value="请填写二级单位商品数量（如需拆零）")
            ws.cell(row=1, column=7, value="请填写最小单位商品数量")
            ws.cell(row=1, column=8, value="单位")

            row_idx = 2
            for store_name, products in data.items():
                for code, qty, remark in products:
                    ws.cell(row=row_idx, column=1, value=store_name)
                    ws.cell(row=row_idx, column=2, value=f"商品_{code}")
                    ws.cell(row=row_idx, column=3, value=code)
                    ws.cell(row=row_idx, column=4, value=remark)
                    ws.cell(row=row_idx, column=5, value="标规")
                    ws.cell(row=row_idx, column=6, value=qty)
                    ws.cell(row=row_idx, column=7, value="")
                    ws.cell(row=row_idx, column=8, value="盒")
                    row_idx += 1

            wb.save(str(file_path))
            wb.close()
            return file_path
        yield _make

    def test_format_b_assigns_order_numbers(self, isolated_hlmc_parser, make_format_b_workbook):
        data = {
            "银泰店": [("GM01", 10, ""), ("GM02", 5, "")],
            "金桥店": [("GM01", 20, "")],
        }
        fpath = make_format_b_workbook(data)

        _, items = hlmc_parser.parse_hlmc_excel(str(fpath))
        assert len(items) >= 2

        store_orders = {}
        for item in items:
            org = item["receiver_org"]
            no = item["order_no"]
            store_orders.setdefault(org, set()).add(no)

        for org, nos in store_orders.items():
            assert len(nos) == 1
            assert next(iter(nos)).endswith("0001")

    def test_format_b_external_code_column(self, isolated_hlmc_parser, make_format_b_workbook):
        data = {"门店A": [("ECODE01", 8, "加急")]}
        fpath = make_format_b_workbook(data)

        _, items = hlmc_parser.parse_hlmc_excel(str(fpath))
        assert items[0]["item_code"] == "ECODE01"
        assert items[0]["remark"] == "加急"

    def test_format_b_reuse_and_increment(self, isolated_hlmc_parser, make_format_b_workbook):
        fpath1 = make_format_b_workbook({"门店B": [("X01", 10, "")]}, "fmt_b_1.xlsx")
        _, items1 = hlmc_parser.parse_hlmc_excel(str(fpath1))
        no1 = items1[0]["order_no"]

        fpath2 = make_format_b_workbook({"门店B": [("X02", 20, "")]}, "fmt_b_2.xlsx")
        _, items2 = hlmc_parser.parse_hlmc_excel(str(fpath2))
        no2 = items2[0]["order_no"]

        assert no1.endswith("0001")
        assert no2.endswith("0002")
        assert no1 != no2
