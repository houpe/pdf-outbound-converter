import sys
from pathlib import Path

import openpyxl
import pytest

backend_dir = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(backend_dir))

import config
from parsers.excel_parser_hlmc import (
    parse_hlmc_excel, _find_col, _match_store, _parse_format_a, _parse_format_b, _get_val,
)


class TestFindCol:

    def test_exact_match(self):
        headers = {"SKU名称": 1, "外部商品编码": 2}
        assert _find_col(headers, "SKU名称") == 1

    def test_partial_match(self):
        headers = {"请填写商品编码": 3, "下单后结余": 6}
        assert _find_col(headers, "商品编码") == 3


class TestMatchStore:

    def test_yintai(self):
        result = _match_store("银泰店")
        assert result == config.HLMC_RECEIVERS["银泰"]

    def test_jinyintan(self):
        result = _match_store("金银潭分店")
        assert result == config.HLMC_RECEIVERS["金银潭"]

    def test_unknown(self):
        result = _match_store("未知门店XYZ")
        assert result == {}

    def test_jinqiao(self):
        result = _match_store("金桥永旺欢乐牧场")
        assert result == config.HLMC_RECEIVERS["金桥"]


class TestFormatBEdgeCases:

    def _make_format_b_wb(self, headers_and_rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        for r_idx, row in enumerate(headers_and_rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        return wb

    def test_starts_from_row_3_when_extra_header(self, tmp_path):
        wb = self._make_format_b_wb([
            ["门店名称", "SKU名称", "外部商品编码", "请填写二级单位商品数量", "请填写最小单位商品数量", "单位", "规格", "备注"],
            ["请填写...", "", "", "", "", "", "", ""],
            ["银泰店", "牛肉片", "SKU01", 10, "", "盒", "500g", ""],
        ])
        fpath = tmp_path / "fmt_b_extra.xlsx"
        wb.save(str(fpath))
        wb.close()

        headers = {}
        ws2 = openpyxl.load_workbook(str(fpath)).active
        for c in range(1, ws2.max_column + 1):
            v = str(ws2.cell(row=1, column=c).value or "").strip()
            if v:
                headers[v] = c
        ws2.parent.close()

        rows = _parse_format_b(ws2, headers)
        assert "银泰店" in rows


class TestFormatAEdgeCases:

    def _make_format_a_wb(self, tmp_path, data):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="SKU名称")
        ws.cell(row=1, column=2, value="外部商品编码")
        ws.cell(row=1, column=3, value="冻结数量的总和")
        ws.cell(row=1, column=4, value="单位")
        c = 5
        for store in data["stores"]:
            ws.cell(row=1, column=c, value=store)
            c += 1
        ws.cell(row=1, column=c, value="下单后结余")
        surplus_col = c

        for r_idx, (sku, code, store_qtys) in enumerate(data["items"], start=2):
            ws.cell(row=r_idx, column=1, value=sku)
            ws.cell(row=r_idx, column=2, value=code)
            ws.cell(row=r_idx, column=3, value=0)
            ws.cell(row=r_idx, column=4, value="盒")
            c = 5
            for qty in store_qtys:
                ws.cell(row=r_idx, column=c, value=qty)
                c += 1
            ws.cell(row=r_idx, column=surplus_col, value=100)

        wb.save(str(tmp_path / "fmt_a.xlsx"))
        wb.close()
        return tmp_path / "fmt_a.xlsx"

    def test_zero_qty_skipped(self, tmp_path):
        fpath = self._make_format_a_wb(tmp_path, {
            "stores": ["银泰"],
            "items": [
                ("牛肉片", "SKU01", [0]),
                ("羊肉卷", "SKU02", [5]),
            ]
        })
        _, items = parse_hlmc_excel(str(fpath))
        codes = {i["item_code"] for i in items}
        assert "SKU02" in codes
        assert "SKU01" not in codes

    def test_none_qty_skipped(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="SKU名称")
        ws.cell(row=1, column=2, value="外部商品编码")
        ws.cell(row=1, column=3, value="冻结数量的总和")
        ws.cell(row=1, column=4, value="单位")
        ws.cell(row=1, column=5, value="银泰")
        ws.cell(row=1, column=6, value="下单后结余")
        ws.cell(row=2, column=1, value="牛肉片")
        ws.cell(row=2, column=2, value="SKU01")
        ws.cell(row=2, column=3, value=0)
        ws.cell(row=2, column=4, value="盒")
        ws.cell(row=2, column=5, value=None)
        ws.cell(row=2, column=6, value=100)
        wb.save(str(tmp_path / "none_qty.xlsx"))
        wb.close()

        _, items = parse_hlmc_excel(str(tmp_path / "none_qty.xlsx"))
        assert len(items) == 0


class TestFormatBIntegration:

    def test_format_b_parses_correctly(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="门店名称")
        ws.cell(row=1, column=2, value="SKU名称")
        ws.cell(row=1, column=3, value="外部商品编码")
        ws.cell(row=1, column=4, value="备注")
        ws.cell(row=1, column=5, value="规格")
        ws.cell(row=1, column=6, value="请填写二级单位商品数量")
        ws.cell(row=1, column=7, value="请填写最小单位商品数量")
        ws.cell(row=1, column=8, value="单位")

        ws.cell(row=2, column=1, value="银泰店")
        ws.cell(row=2, column=2, value="商品A")
        ws.cell(row=2, column=3, value="CODE01")
        ws.cell(row=2, column=6, value=10)
        ws.cell(row=2, column=8, value="盒")

        ws.cell(row=3, column=1, value="金桥店")
        ws.cell(row=3, column=2, value="商品B")
        ws.cell(row=3, column=3, value="CODE02")
        ws.cell(row=3, column=7, value=20)  # Only min qty
        ws.cell(row=3, column=8, value="盒")

        fpath = tmp_path / "fmt_b_int.xlsx"
        wb.save(str(fpath))
        wb.close()

        _, items = parse_hlmc_excel(str(fpath))
        assert len(items) == 2
        codes = {i["item_code"] for i in items}
        assert codes == {"CODE01", "CODE02"}

    def test_both_secondary_and_min_prefers_secondary(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="门店名称")
        ws.cell(row=1, column=2, value="SKU名称")
        ws.cell(row=1, column=3, value="外部商品编码")
        ws.cell(row=1, column=4, value="备注")
        ws.cell(row=1, column=5, value="规格")
        ws.cell(row=1, column=6, value="请填写二级单位商品数量")
        ws.cell(row=1, column=7, value="请填写最小单位商品数量")
        ws.cell(row=1, column=8, value="单位")

        ws.cell(row=2, column=1, value="银泰店")
        ws.cell(row=2, column=2, value="商品A")
        ws.cell(row=2, column=3, value="CODE01")
        ws.cell(row=2, column=6, value=5)   # secondary
        ws.cell(row=2, column=7, value=10)  # min
        ws.cell(row=2, column=8, value="盒")

        fpath = tmp_path / "fmt_b_both.xlsx"
        wb.save(str(fpath))
        wb.close()

        _, items = parse_hlmc_excel(str(fpath))
        assert items[0]["quantity"] == "5"

    def test_float_qty_converted_to_int(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="门店名称")
        ws.cell(row=1, column=2, value="SKU名称")
        ws.cell(row=1, column=3, value="外部商品编码")
        ws.cell(row=1, column=4, value="备注")
        ws.cell(row=1, column=5, value="规格")
        ws.cell(row=1, column=6, value="请填写二级单位商品数量")
        ws.cell(row=1, column=7, value="请填写最小单位商品数量")
        ws.cell(row=1, column=8, value="单位")

        ws.cell(row=2, column=1, value="门店")
        ws.cell(row=2, column=2, value="商品")
        ws.cell(row=2, column=3, value="C1")
        ws.cell(row=2, column=6, value=10.7)
        ws.cell(row=2, column=8, value="盒")

        fpath = tmp_path / "float_qty.xlsx"
        wb.save(str(fpath))
        wb.close()

        _, items = parse_hlmc_excel(str(fpath))
        assert items[0]["quantity"] == "10"
