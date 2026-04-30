"""
Tests for parser functions in parsers module.
"""

import os
import sys
from pathlib import Path

import openpyxl
import pytest

# Add backend to path
backend_dir = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(backend_dir))

# Import from parsers module
from parsers import (
    _extract_shop_name,
    _extract_header_value,
    parse_lmt_excel,
    parse_hlmc_excel,
)


class TestExtractShopName:
    """Tests for _extract_shop_name function."""

    def test_extract_with_date_prefix(self):
        """Test extraction from filename with date prefix like '12.25门店名-xxx'."""
        result = _extract_shop_name("12.25测试门店-配送发货单PS2512210002001.xlsx")
        assert result == "测试门店"

    def test_extract_with_dot_separator(self):
        """Test extraction with dot in date prefix."""
        result = _extract_shop_name("12.25.测试门店-订单.xlsx")
        # The function removes leading digits and dots
        assert "测试门店" in result or result.startswith("测试")

    def test_extract_with_em_dash(self):
        """Test extraction with em dash separator."""
        result = _extract_shop_name("2024.01.15门店名称—发货单.xlsx")
        assert "门店名称" in result or result == "门店名称"

    def test_extract_simple_filename(self):
        """Test extraction from filename without date prefix."""
        result = _extract_shop_name("门店名称.xlsx")
        assert result == "门店名称"

    def test_extract_with_parentheses(self):
        """Test extraction with parentheses in filename."""
        result = _extract_shop_name("12.25门店A(分店).xlsx")
        assert "门店A" in result

    def test_extract_with_chinese_parentheses(self):
        """Test extraction with Chinese parentheses."""
        result = _extract_shop_name("12.25门店B（分店）.xlsx")
        assert "门店B" in result


class TestExtractHeaderValue:
    """Tests for _extract_header_value function."""

    def test_extract_single_value(self):
        """Test extracting a single header value."""
        text = "单据编号：PS2512210002001\n收货机构：测试门店"
        result = _extract_header_value(text, "单据编号")
        assert result == "PS2512210002001"

    def test_extract_with_colon(self):
        """Test extraction with Chinese colon."""
        text = "收货机构:测试门店A\n供货机构：中央厨房"
        result = _extract_header_value(text, "收货机构")
        assert result == "测试门店A"

    def test_extract_missing_label(self):
        """Test extraction when label is not present."""
        text = "单据编号：PS2512210002001\n收货机构：测试门店"
        result = _extract_header_value(text, "不存在字段")
        assert result == ""

    def test_extract_multiline_value(self):
        """Test extraction with multiline text."""
        text = "单据编号：PS2512210002001\n收货机构：测试门店\n供货机构：中央厨房"
        result = _extract_header_value(text, "收货机构")
        assert result == "测试门店"

    def test_extract_with_spaces(self):
        """Test extraction with spaces around the value."""
        text = "单据编号：  PS2512210002001  \n收货机构：测试门店"
        result = _extract_header_value(text, "单据编号")
        assert result == "PS2512210002001"


class TestParseLmtExcel:
    """Tests for parse_lmt_excel function."""

    def test_parse_basic_lmt(self, lmt_excel: Path):
        """Test parsing a basic LMT Excel file."""
        header, items = parse_lmt_excel(str(lmt_excel), "12.25测试门店-配送发货单PS2512210002001.xlsx")

        assert header["receiver_org"] == "测试门店A"
        assert header["supplier_org"] == "中央厨房"
        assert header["order_no"] == "PS2512210002001"
        assert header["receiver_name"] == "张三"
        assert header["receiver_phone"] == "13800138000"
        assert header["receiver_address"] == "北京市朝阳区测试路1号"

        # Should parse items from row 2 (first item) and row 10 (second item)
        assert len(items) >= 1
        assert items[0]["item_code"] == "CODE001"
        assert items[0]["item_name"] == "测试商品A"
        assert items[0]["quantity"] == "10"

    def test_parse_lmt_no_filename(self, lmt_excel: Path):
        """Test parsing LMT Excel without filename."""
        header, items = parse_lmt_excel(str(lmt_excel))

        assert header["receiver_org"] == "测试门店A"
        assert len(items) >= 1

    def test_parse_lmt_missing_columns(self, test_files_dir: Path):
        """Test parsing LMT Excel with minimal columns."""
        file_path = test_files_dir / "lmt_minimal.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "配送发货单"

        # Minimal data - item header in row 1 (within parser's search range 1-6)
        ws.cell(row=1, column=3, value="物品编码")
        ws.cell(row=1, column=4, value="物品名称")
        ws.cell(row=1, column=7, value="发货数量")
        ws.cell(row=3, column=1, value="收货机构")
        ws.cell(row=3, column=2, value="简化门店")

        # Item data in row 2
        ws.cell(row=2, column=3, value="CODE003")
        ws.cell(row=2, column=4, value="简化商品")
        ws.cell(row=2, column=7, value=3)

        wb.save(str(file_path))
        wb.close()

        header, items = parse_lmt_excel(str(file_path))

        assert header["receiver_org"] == "简化门店"
        assert len(items) == 1
        assert items[0]["item_code"] == "CODE003"

    def test_parse_lmt_empty_file(self, test_files_dir: Path):
        """Test parsing an empty LMT Excel file."""
        file_path = test_files_dir / "lmt_empty.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        wb.save(str(file_path))
        wb.close()

        header, items = parse_lmt_excel(str(file_path))

        # Should return empty results without crashing
        assert isinstance(header, dict)
        assert isinstance(items, list)


class TestParseHlmcExcel:
    """Tests for parse_hlmc_excel function."""

    def test_parse_basic_hlmc(self, hlmc_excel: Path):
        """Test parsing a basic HLMC Excel file."""
        header, items = parse_hlmc_excel(str(hlmc_excel))

        # Header should be empty for HLMC (info is per-item)
        assert isinstance(header, dict)

        # Should have items for each store with quantity > 0
        assert len(items) >= 2

        # Check first item structure
        first_item = items[0]
        assert "item_code" in first_item
        assert "item_name" in first_item
        assert "quantity" in first_item
        assert "receiver_org" in first_item

    def test_parse_hlmc_store_matching(self, hlmc_excel: Path):
        """Test that HLMC parser correctly matches store names to receivers."""
        header, items = parse_hlmc_excel(str(hlmc_excel))

        # Check that items have store names
        stores = {item["receiver_org"] for item in items}
        assert "银泰" in stores or "金银潭" in stores

    def test_parse_hlmc_missing_surplus_column(self, test_files_dir: Path):
        """Test that parsing fails without required '下单后结余' column."""
        file_path = test_files_dir / "hlmc_invalid.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="SKU名称")
        ws.cell(row=1, column=2, value="外部商品编码")
        # Missing '下单后结余' column

        wb.save(str(file_path))
        wb.close()

        with pytest.raises(ValueError, match="下单后结余"):
            parse_hlmc_excel(str(file_path))

    def test_parse_hlmc_zero_quantities(self, test_files_dir: Path):
        """Test parsing HLMC file where all quantities are zero."""
        file_path = test_files_dir / "hlmc_zero.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="SKU名称")
        ws.cell(row=1, column=2, value="外部商品编码")
        ws.cell(row=1, column=3, value="冻结数量的总和")
        ws.cell(row=1, column=4, value="单位")
        ws.cell(row=1, column=5, value="银泰")
        ws.cell(row=1, column=6, value="下单后结余")

        # Zero quantities
        ws.cell(row=2, column=1, value="牛肉片")
        ws.cell(row=2, column=2, value="SKU001")
        ws.cell(row=2, column=3, value=0)
        ws.cell(row=2, column=4, value="盒")
        ws.cell(row=2, column=5, value=0)  # Zero quantity
        ws.cell(row=2, column=6, value=100)

        wb.save(str(file_path))
        wb.close()

        header, items = parse_hlmc_excel(str(file_path))

        # Should return empty items (no quantity > 0)
        assert len(items) == 0

    def test_parse_hlmc_quantity_conversion(self, test_files_dir: Path):
        """Test that HLMC parser correctly converts quantities to integers."""
        file_path = test_files_dir / "hlmc_float.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="SKU名称")
        ws.cell(row=1, column=2, value="外部商品编码")
        ws.cell(row=1, column=3, value="冻结数量的总和")
        ws.cell(row=1, column=4, value="单位")
        ws.cell(row=1, column=5, value="银泰")
        ws.cell(row=1, column=6, value="下单后结余")

        ws.cell(row=2, column=1, value="牛肉片")
        ws.cell(row=2, column=2, value="SKU001")
        ws.cell(row=2, column=3, value=0)
        ws.cell(row=2, column=4, value="盒")
        ws.cell(row=2, column=5, value=10.5)  # Float quantity
        ws.cell(row=2, column=6, value=100)

        wb.save(str(file_path))
        wb.close()

        header, items = parse_hlmc_excel(str(file_path))

        assert len(items) == 1
        # Quantity should be converted to integer
        assert items[0]["quantity"] == "10"