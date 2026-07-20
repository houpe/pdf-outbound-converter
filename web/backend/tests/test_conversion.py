import sys
from pathlib import Path

import openpyxl
import pytest
from fastapi import HTTPException

backend_dir = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(backend_dir))

import database
from services.conversion import create_excel


@pytest.fixture
def db_conn(tmp_path: Path, monkeypatch):
    db_path = tmp_path / "conv.db"
    conn = database.sqlite3.connect(str(db_path))
    conn.execute("""
        CREATE TABLE IF NOT EXISTS split_codes (
            code TEXT NOT NULL,
            split TEXT NOT NULL DEFAULT '是',
            item_name TEXT,
            warehouse_code TEXT NOT NULL DEFAULT 'ZTOWHHY001',
            created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')),
            UNIQUE(code COLLATE NOCASE, warehouse_code)
        )
    """)
    conn.commit()
    conn.close()

    def mock_get_db():
        c = database.sqlite3.connect(str(db_path))
        c.row_factory = database.sqlite3.Row
        return c

    monkeypatch.setattr(database, "get_db", mock_get_db)
    monkeypatch.setattr(database, "DB_PATH", db_path)

    conn = mock_get_db()
    conn.execute("INSERT INTO split_codes (code, split, warehouse_code) VALUES ('C1', '是', 'ZTOWHHY001')")
    conn.execute("INSERT INTO split_codes (code, split, warehouse_code) VALUES ('B1', '是', 'ZTOWHHY001')")
    conn.execute("INSERT INTO split_codes (code, split, warehouse_code) VALUES ('N', '是', 'ZTOWHHY001')")
    conn.commit()
    conn.close()
    return db_path


@pytest.fixture
def sample_oms_template(tmp_path: Path) -> Path:
    p = tmp_path / "oms.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    values = ["单据编号", "商户编码", "仓库编码", "备注", "收货人信息", "收货机构", "商品名称", "商品编码", "二级单位数量", "最小单位数量"]
    for col, val in enumerate(values, start=1):
        ws.cell(row=1, column=col, value=val)
    wb.save(str(p))
    wb.close()
    return p


class TestCreateExcel:

    def test_basic_write(self, sample_oms_template, db_conn, tmp_path):
        out_path = str(tmp_path / "r.xlsx")
        items = [{"order_no": "PS001", "receiver_org": "门店", "receiver_name": "张", "receiver_phone": "138", "receiver_address": "京", "item_name": "N", "item_code": "C1", "quantity": "10"}]
        create_excel({}, items, str(sample_oms_template), out_path, merchant_code="M001", template_key="qzz")
        wb = openpyxl.load_workbook(out_path)
        assert wb.active.cell(row=3, column=1).value == "PS001"
        assert wb.active.cell(row=3, column=2).value == "M001"
        assert wb.active.cell(row=3, column=3).value == "ZTOWHHY001"
        assert wb.active.cell(row=3, column=6).value == "门店"
        assert wb.active.cell(row=3, column=7).value == "N"
        assert wb.active.cell(row=3, column=8).value == "C1"
        assert wb.active.cell(row=3, column=9).value is None
        assert wb.active.cell(row=3, column=10).value == 10
        wb.close()

    def test_qzz_routes_to_col10(self, sample_oms_template, db_conn, tmp_path):
        out_path = str(tmp_path / "qzz.xlsx")
        items = [{"order_no": "O1", "receiver_org": "S1", "receiver_name": "", "receiver_phone": "", "receiver_address": "", "item_name": "N", "item_code": "C1", "quantity": "5"}]
        create_excel({}, items, str(sample_oms_template), out_path, merchant_code="MC", template_key="qzz")
        assert openpyxl.load_workbook(out_path).active.cell(row=3, column=10).value == 5

    def test_split_yes_routes_to_col9(self, sample_oms_template, db_conn, tmp_path):
        conn = database.get_db()
        conn.execute("INSERT OR REPLACE INTO split_codes (code, split, warehouse_code) VALUES ('C1', '是', 'ZTOWHHY001')")
        conn.commit()
        conn.close()

        out_path = str(tmp_path / "sy.xlsx")
        items = [{"order_no": "O1", "receiver_org": "S1", "receiver_name": "", "receiver_phone": "", "receiver_address": "", "item_name": "N", "item_code": "C1", "quantity": "7"}]
        create_excel({}, items, str(sample_oms_template), out_path, merchant_code="MC", template_key="lmt")
        wb = openpyxl.load_workbook(out_path)
        assert wb.active.cell(row=3, column=9).value == 7
        assert wb.active.cell(row=3, column=10).value is None

    def test_split_no_routes_to_col10(self, sample_oms_template, db_conn, tmp_path):
        conn = database.get_db()
        conn.execute("INSERT OR REPLACE INTO split_codes (code, split, warehouse_code) VALUES ('C1', '否', 'ZTOWHHY001')")
        conn.commit()
        conn.close()

        out_path = str(tmp_path / "sn.xlsx")
        items = [{"order_no": "O1", "receiver_org": "S1", "receiver_name": "", "receiver_phone": "", "receiver_address": "", "item_name": "N", "item_code": "C1", "quantity": "7"}]
        create_excel({}, items, str(sample_oms_template), out_path, merchant_code="MC", template_key="lmt")
        wb = openpyxl.load_workbook(out_path)
        assert wb.active.cell(row=3, column=9).value is None
        assert wb.active.cell(row=3, column=10).value == 7

    def test_sorts_by_store(self, sample_oms_template, db_conn, tmp_path):
        out_path = str(tmp_path / "srt.xlsx")
        items = [
            {"order_no": "", "receiver_org": "店B", "receiver_name": "", "receiver_phone": "", "receiver_address": "", "item_name": "B1", "item_code": "B1", "quantity": "1"},
            {"order_no": "", "receiver_org": "店A", "receiver_name": "", "receiver_phone": "", "receiver_address": "", "item_name": "A1", "item_code": "A1", "quantity": "1"},
        ]
        create_excel({}, items, str(sample_oms_template), out_path, merchant_code="MC", template_key="qzz")
        wb = openpyxl.load_workbook(out_path)
        assert wb.active.cell(row=3, column=6).value == "店A"
        assert wb.active.cell(row=4, column=6).value == "店B"

    def test_lmt_validates_missing_splits(self, sample_oms_template, db_conn, tmp_path):
        out_path = str(tmp_path / "fail.xlsx")
        items = [{"order_no": "", "receiver_org": "S1", "receiver_name": "", "receiver_phone": "", "receiver_address": "", "item_name": "N", "item_code": "MISSING", "quantity": "1"}]
        with pytest.raises(HTTPException) as exc:
            create_excel({}, items, str(sample_oms_template), out_path, merchant_code="MC", template_key="lmt")
        assert exc.value.status_code == 400
        assert "商品编码缺失" in exc.value.detail["error"]

    def test_non_lmt_skips_validation(self, sample_oms_template, db_conn, tmp_path):
        out_path = str(tmp_path / "skip.xlsx")
        items = [{"order_no": "", "receiver_org": "S1", "receiver_name": "", "receiver_phone": "", "receiver_address": "", "item_name": "N", "item_code": "MISSING", "quantity": "1"}]
        create_excel({}, items, str(sample_oms_template), out_path, merchant_code="MC", template_key="qzz")
        assert openpyxl.load_workbook(out_path).active.cell(row=3, column=8).value == "MISSING"

    def test_handles_float_quantity(self, sample_oms_template, db_conn, tmp_path):
        out_path = str(tmp_path / "fl.xlsx")
        items = [{"order_no": "O1", "receiver_org": "S1", "receiver_name": "", "receiver_phone": "", "receiver_address": "", "item_name": "N", "item_code": "C1", "quantity": "12.7"}]
        create_excel({}, items, str(sample_oms_template), out_path, merchant_code="MC", template_key="qzz")
        assert openpyxl.load_workbook(out_path).active.cell(row=3, column=10).value == 12

    def test_contact_concatenated(self, sample_oms_template, db_conn, tmp_path):
        out_path = str(tmp_path / "ct.xlsx")
        items = [{"order_no": "", "receiver_org": "S1", "receiver_name": "张", "receiver_phone": "138", "receiver_address": "京", "item_name": "N", "item_code": "C1", "quantity": "1"}]
        create_excel({}, items, str(sample_oms_template), out_path, merchant_code="MC", template_key="qzz")
        # 收件人名字统一加“收”字后缀：张 → 张收
        assert openpyxl.load_workbook(out_path).active.cell(row=3, column=5).value == "张收,138,京"
