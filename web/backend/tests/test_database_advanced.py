import sys
from pathlib import Path

import openpyxl
import pytest

backend_dir = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(backend_dir))

import database
import config


class TestGetSplitMap:

    def test_empty_returns_empty_dict(self, tmp_path, monkeypatch):
        db = tmp_path / "sm.db"
        conn = database.sqlite3.connect(str(db))
        conn.execute("CREATE TABLE IF NOT EXISTS split_codes (code TEXT PRIMARY KEY COLLATE NOCASE, split TEXT NOT NULL DEFAULT '是', item_name TEXT, created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')))")
        conn.commit()
        conn.close()
        monkeypatch.setattr(database, "DB_PATH", db)
        assert database.get_split_map() == {}

    def test_returns_mapping(self, tmp_path, monkeypatch):
        db = tmp_path / "sm2.db"
        conn = database.sqlite3.connect(str(db))
        conn.execute("CREATE TABLE IF NOT EXISTS split_codes (code TEXT PRIMARY KEY COLLATE NOCASE, split TEXT NOT NULL DEFAULT '是', item_name TEXT, created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')))")
        conn.execute("INSERT INTO split_codes (code, split) VALUES ('ABC123', '是')")
        conn.execute("INSERT INTO split_codes (code, split) VALUES ('DEF456', '否')")
        conn.commit()
        conn.close()
        monkeypatch.setattr(database, "DB_PATH", db)
        result = database.get_split_map()
        assert result["abc123"] == "是"
        assert result["def456"] == "否"

    def test_case_lower_in_map(self, tmp_path, monkeypatch):
        db = tmp_path / "sm3.db"
        conn = database.sqlite3.connect(str(db))
        conn.execute("CREATE TABLE IF NOT EXISTS split_codes (code TEXT PRIMARY KEY COLLATE NOCASE, split TEXT NOT NULL DEFAULT '是', item_name TEXT, created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')))")
        conn.execute("INSERT INTO split_codes (code, split) VALUES ('UPCASE', '是')")
        conn.commit()
        conn.close()
        monkeypatch.setattr(database, "DB_PATH", db)
        result = database.get_split_map()
        assert "upcase" in result


class TestGetDb:

    def test_returns_connection_with_row_factory(self, tmp_path, monkeypatch):
        db = tmp_path / "getdb.db"
        conn = database.sqlite3.connect(str(db))
        conn.execute("CREATE TABLE test (id INTEGER PRIMARY KEY)")
        conn.commit()
        conn.close()
        monkeypatch.setattr(database, "DB_PATH", db)
        c = database.get_db()
        assert isinstance(c, database.sqlite3.Connection)
        c.execute("SELECT 1")
        c.close()


class TestInitDb:

    def test_creates_all_tables(self, tmp_path, monkeypatch):
        db = tmp_path / "init_all.db"
        monkeypatch.setattr(database, "DB_PATH", db)
        monkeypatch.setattr(database, "SPLIT_TEMPLATE", Path("/nonexistent"))

        database.init_db()

        conn = database.sqlite3.connect(str(db))
        tables = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
        assert "split_codes" in tables
        assert "hlmc_sequences" in tables
        assert "hlmc_history" in tables
        conn.close()

    def test_creates_index(self, tmp_path, monkeypatch):
        db = tmp_path / "init_idx.db"
        monkeypatch.setattr(database, "DB_PATH", db)
        monkeypatch.setattr(database, "SPLIT_TEMPLATE", Path("/nonexistent"))

        database.init_db()

        conn = database.sqlite3.connect(str(db))
        indexes = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='index' AND name LIKE 'idx_%'").fetchall()]
        assert "idx_hlmc_hist_sig" in indexes
        conn.close()

    def test_added_columns_exist(self, tmp_path, monkeypatch):
        db = tmp_path / "init_cols.db"
        monkeypatch.setattr(database, "DB_PATH", db)
        monkeypatch.setattr(database, "SPLIT_TEMPLATE", Path("/nonexistent"))

        database.init_db()

        conn = database.sqlite3.connect(str(db))
        cols = {r[1] for r in conn.execute("PRAGMA table_info(split_codes)").fetchall()}
        assert "created_at" in cols
        assert "item_name" in cols
        conn.close()


class TestSeedSplitCodes:

    def _make_template(self, tmp_path, rows, headers=None):
        if headers is None:
            headers = ["商品编码", "是否拆零"]
        tpl = tmp_path / "seed_tpl.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        for c, h in enumerate(headers, start=1):
            ws.cell(row=1, column=c, value=h)
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        wb.save(str(tpl))
        wb.close()
        return tpl

    def test_missing_template_no_error(self, tmp_path, monkeypatch):
        db = tmp_path / "s1.db"
        conn = database.sqlite3.connect(str(db))
        conn.execute("CREATE TABLE IF NOT EXISTS split_codes (code TEXT PRIMARY KEY COLLATE NOCASE, split TEXT NOT NULL DEFAULT '是', item_name TEXT, created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')))")
        conn.commit()

        monkeypatch.setattr(database, "SPLIT_TEMPLATE", Path("/nonexistent/template.xlsx"))
        database.seed_split_codes(conn)  # no exception
        assert conn.execute("SELECT COUNT(*) FROM split_codes").fetchone()[0] == 0
        conn.close()

    def test_template_valid_data(self, tmp_path, monkeypatch):
        tpl = self._make_template(tmp_path, [
            ["SEED01", "是"],
            ["SEED02", "否"],
        ])
        monkeypatch.setattr(database, "SPLIT_TEMPLATE", tpl)

        db = tmp_path / "s2.db"
        conn = database.sqlite3.connect(str(db))
        conn.execute("CREATE TABLE IF NOT EXISTS split_codes (code TEXT PRIMARY KEY COLLATE NOCASE, split TEXT NOT NULL DEFAULT '是', item_name TEXT, created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')))")
        conn.commit()

        database.seed_split_codes(conn)
        count = conn.execute("SELECT COUNT(*) FROM split_codes").fetchone()[0]
        assert count == 2
        conn.close()

    def test_missing_columns_skipped(self, tmp_path, monkeypatch):
        tpl = self._make_template(tmp_path, [["X"]], headers=["wrong_column"])
        monkeypatch.setattr(database, "SPLIT_TEMPLATE", tpl)

        db = tmp_path / "s3.db"
        conn = database.sqlite3.connect(str(db))
        conn.execute("CREATE TABLE IF NOT EXISTS split_codes (code TEXT PRIMARY KEY COLLATE NOCASE, split TEXT NOT NULL DEFAULT '是', item_name TEXT, created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')))")
        conn.commit()

        database.seed_split_codes(conn)
        count = conn.execute("SELECT COUNT(*) FROM split_codes").fetchone()[0]
        assert count == 0
        conn.close()

    def test_empty_code_skipped(self, tmp_path, monkeypatch):
        tpl = self._make_template(tmp_path, [["", "是"]])
        monkeypatch.setattr(database, "SPLIT_TEMPLATE", tpl)

        db = tmp_path / "s4.db"
        conn = database.sqlite3.connect(str(db))
        conn.execute("CREATE TABLE IF NOT EXISTS split_codes (code TEXT PRIMARY KEY COLLATE NOCASE, split TEXT NOT NULL DEFAULT '是', item_name TEXT, created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')))")
        conn.commit()

        database.seed_split_codes(conn)
        count = conn.execute("SELECT COUNT(*) FROM split_codes").fetchone()[0]
        assert count == 0
        conn.close()

    def test_invalid_split_defaults_to_yes(self, tmp_path, monkeypatch):
        tpl = self._make_template(tmp_path, [["BADVAL", "maybe"]])
        monkeypatch.setattr(database, "SPLIT_TEMPLATE", tpl)

        db = tmp_path / "s5.db"
        conn = database.sqlite3.connect(str(db))
        conn.execute("CREATE TABLE IF NOT EXISTS split_codes (code TEXT PRIMARY KEY COLLATE NOCASE, split TEXT NOT NULL DEFAULT '是', item_name TEXT, created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')))")
        conn.commit()

        database.seed_split_codes(conn)
        row = conn.execute("SELECT split FROM split_codes WHERE code = 'BADVAL'").fetchone()
        assert row[0] == "是"
        conn.close()

    def test_does_not_overwrite_existing(self, tmp_path, monkeypatch):
        tpl = self._make_template(tmp_path, [["EXISTING", "否"]])
        monkeypatch.setattr(database, "SPLIT_TEMPLATE", tpl)

        db = tmp_path / "s6.db"
        conn = database.sqlite3.connect(str(db))
        conn.execute("CREATE TABLE IF NOT EXISTS split_codes (code TEXT PRIMARY KEY COLLATE NOCASE, split TEXT NOT NULL DEFAULT '是', item_name TEXT, created_at TEXT NOT NULL DEFAULT (datetime('now', 'localtime')))")
        conn.execute("INSERT INTO split_codes (code, split) VALUES ('EXISTING', '是')")
        conn.commit()

        database.seed_split_codes(conn)
        row = conn.execute("SELECT split FROM split_codes WHERE code = 'EXISTING'").fetchone()
        assert row[0] == "是"
        conn.close()
