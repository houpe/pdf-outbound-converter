#!/usr/bin/env python3.10
"""
PDF出库单转Excel工具
读取PDF配送单，转换为OMS出库Excel格式
"""

import pdfplumber
import openpyxl
import re
import os
import sys


def extract_pdf_data(pdf_path):
    """从PDF中提取出库单数据"""
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        all_tables = []
        
        for page in pdf.pages:
            full_text += page.extract_text() or ""
            tables = page.extract_tables()
            all_tables.extend(tables)
    
    header_info = parse_header(full_text)
    items = parse_items(all_tables)
    
    return header_info, items


def resource_path(relative_path):
    """Return the absolute path to a resource, works for both normal and PyInstaller bundled apps."""
    if getattr(sys, 'frozen', False):
        base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)


def parse_header(text):
    """解析PDF头部信息"""
    info = {}
    
    match = re.search(r'单据编号[：:]\s*(\S+)', text)
    info['order_no'] = match.group(1) if match else ''
    
    match = re.search(r'收货机构[：:]\s*([^  ]+?)(?=\s+[^\s：:]+[：:]|$)', text)
    info['receiver_org'] = match.group(1).strip() if match else ''
    
    match = re.search(r'供货机构[：:]\s*([^  ]+?)(?=\s+[^\s：:]+[：:]|$)', text)
    info['supplier_org'] = match.group(1).strip() if match else ''
    
    match = re.search(r'收货人[：:]\s*([^  ]+?)(?=\s+[^\s：:]+[：:]|$)', text)
    info['receiver_name'] = match.group(1).strip() if match else ''
    
    match = re.search(r'收货电话[：:]\s*(\d+)', text)
    info['receiver_phone'] = match.group(1) if match else ''
    
    match = re.search(r'收货地址[：:]\s*(.+?)(?=\n|$)', text)
    info['receiver_address'] = match.group(1).strip() if match else ''
    
    match = re.search(r'订单日期[：:]\s*(\S+)', text)
    info['order_date'] = match.group(1) if match else ''
    
    return info


def parse_items(tables):
    """解析商品明细表格"""
    items = []
    
    for table in tables:
        for row in table:
            if not row or len(row) < 6:
                continue
            
            first_cell = str(row[0]).strip()
            if not first_cell.isdigit():
                continue
            
            item = {
                'category': str(row[1]).strip() if row[1] else '',
                'item_code': str(row[2]).strip() if row[2] else '',
                'item_name': str(row[3]).strip() if row[3] else '',
                'spec': str(row[4]).strip().replace('\n', '') if row[4] else '',
                'unit': str(row[5]).strip() if row[5] else '',
                'quantity': str(row[6]).strip() if row[6] else '0',
                'remark': str(row[7]).strip() if row[7] else '',
            }
            
            if item['item_code'] and item['item_name']:
                items.append(item)
    
    return items


def create_excel(header_info, items, template_path, output_path, merchant_code=''):
    """创建OMS出库Excel文件"""
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            cell.value = None
    
    receiver_info = f"{header_info.get('receiver_name', '')},{header_info.get('receiver_phone', '')},{header_info.get('receiver_address', '')}"
    
    for i, item in enumerate(items, start=3):
        ws.cell(row=i, column=1, value=header_info.get('order_no', ''))
        ws.cell(row=i, column=2, value=merchant_code)
        ws.cell(row=i, column=3, value=header_info.get('supplier_org', ''))
        ws.cell(row=i, column=4, value='')
        ws.cell(row=i, column=5, value=receiver_info)
        ws.cell(row=i, column=6, value=item['item_code'])
        ws.cell(row=i, column=7, value='')
        ws.cell(row=i, column=8, value=int(item['quantity']) if item['quantity'].isdigit() else item['quantity'])
        ws.cell(row=i, column=9, value=header_info.get('receiver_org', ''))
        ws.cell(row=i, column=10, value=item['remark'])
    
    wb.save(output_path)
    print(f"Excel已保存至: {output_path}")
    print(f"共处理 {len(items)} 条商品记录")


def main():
    """主函数"""
    # 使用资源路径，以支持打包后的可执行文件
    script_dir = os.path.dirname(os.path.abspath(__file__))
    default_pdf = resource_path('黔寨寨贵州烙锅（鞍山店）冻库.pdf')
    # 将输出放在当前工作目录，避免打包后写入临时目录
    default_output = os.path.join(os.getcwd(), '出库单_转换结果.xlsx')
    template_path = resource_path('OMS出库.xlsx')
    
    if len(sys.argv) > 1:
        pdf_path = sys.argv[1]
    else:
        pdf_path = default_pdf
    
    if len(sys.argv) > 2:
        output_path = sys.argv[2]
    else:
        output_path = default_output
    
    if not os.path.exists(pdf_path):
        print(f"错误: 找不到PDF文件 {pdf_path}")
        sys.exit(1)
    
    if not os.path.exists(template_path):
        print(f"错误: 找不到Excel模板文件 {template_path}")
        sys.exit(1)
    
    print(f"正在读取PDF: {pdf_path}")
    header_info, items = extract_pdf_data(pdf_path)
    
    print("\n=== 提取的头部信息 ===")
    for key, value in header_info.items():
        print(f"  {key}: {value}")
    
    print(f"\n=== 提取的商品明细 ({len(items)}条) ===")
    for item in items:
        print(f"  {item['item_code']} - {item['item_name']} x {item['quantity']} {item['unit']}")
    
    merchant_code = input("\n请输入商户编码: ").strip()
    if not merchant_code:
        print("警告: 商户编码为空，将使用空值")
    
    print("\n=== 正在生成Excel ===")
    create_excel(header_info, items, template_path, output_path, merchant_code)
    print("\n转换完成!")


if __name__ == '__main__':
    main()
